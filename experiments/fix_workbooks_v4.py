"""
Fix workbooks v4 (2026-08-13): add Price Basis Corrected flag columns.

Changes to CORRECTED workbook only:
1. Add column "Price Basis Corrected" (YES/NO) and "Not Corrected Reason"
   to Human_Data_Entry, after the last used column.
2. Apply light orange conditional formatting to Human_Data_Entry rows where
   Price Basis Corrected = NO.
3. Update Human_Data_Entry A1 banner with 294/126 split.
4. Update Corrections_Log with 294/126 summary.
5. Update Accuracy_Conventions: report all human figures again on the
   294-corrected subset only, using the new V column as a filter.

LOCKED workbook is not touched.
"""

import csv, sys
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.formatting.rule import FormulaRule
from openpyxl.utils import get_column_letter

ROOT    = Path(__file__).resolve().parent.parent
WB_DIR  = ROOT / "data" / "workbook"
SUMMARY = ROOT / "outputs" / "global" / "summary"

OUT_CORRECTED = WB_DIR / "Master_Data_CORRECTED_2026-08-13.xlsx"
HUMAN_PRICES  = SUMMARY / "workbook_human_prices_corrected.csv"

BOLD   = Font(bold=True)
ITALIC = Font(italic=True)
GREY   = PatternFill("solid", fgColor="D9D9D9")
YELLOW = PatternFill("solid", fgColor="FFFFC0")
ORANGE = PatternFill("solid", fgColor="FFE0B2")   # light orange for uncorrected rows

N_CORRECTED = 294
N_NOT       = 126
N_TOTAL     = 420

CORRECTED_BANNER = (
    f"HUMAN ARM CORRECTED. Prices M to P re-anchored on release_date from EDGAR filing dates, "
    f"2026-08-13. {N_CORRECTED} of {N_TOTAL} rows re-anchored; {N_NOT} remain on original "
    f"report_date basis (human-only events not in LLM universe — see column 'Price Basis "
    f"Corrected'). Columns Q to U and the Summary, Charts and Efficiency sheets recalculate from "
    f"these. Original basis retained in Master_Data_LOCKED_2026-08-13.xlsx. See Corrections_Log."
)


def load_price_corrections():
    """Return dict {row_index: {'correctable': 'TRUE'/'FALSE', 'reason': str}}."""
    mapping = {}
    with open(HUMAN_PRICES, newline="") as f:
        for row in csv.DictReader(f):
            ri = int(row["row_index"])
            mapping[ri] = {
                "correctable": row["correctable"],
                "reason": row["not_correctable_reason"],
            }
    return mapping


def add_flag_columns(ws_h, corrections):
    """
    Add 'Price Basis Corrected' and 'Not Corrected Reason' columns to Human_Data_Entry.
    Returns (v_col, w_col) as 1-based column indices.
    """
    # Find last used column (scan header row 2, which has column headers)
    last_col = 1
    for c in range(1, 40):
        val = ws_h.cell(2, c).value
        if val is not None:
            last_col = c

    v_col = last_col + 1
    w_col = last_col + 2
    v_letter = get_column_letter(v_col)
    w_letter = get_column_letter(w_col)

    print(f"  Adding flag columns at {v_letter} ({v_col}) and {w_letter} ({w_col})")

    # Write headers in row 2
    hdr_v = ws_h.cell(2, v_col)
    hdr_v.value = "Price Basis Corrected"
    hdr_v.font = BOLD
    hdr_v.fill = GREY

    hdr_w = ws_h.cell(2, w_col)
    hdr_w.value = "Not Corrected Reason"
    hdr_w.font = BOLD
    hdr_w.fill = GREY

    # Write per-row values
    n_yes = 0
    n_no  = 0
    for row_idx, info in corrections.items():
        flag = "YES" if info["correctable"] == "TRUE" else "NO"
        reason = info["reason"] if flag == "NO" else ""
        cell_v = ws_h.cell(row_idx, v_col)
        cell_v.value = flag
        cell_w = ws_h.cell(row_idx, w_col)
        cell_w.value = reason
        if flag == "YES":
            n_yes += 1
        else:
            n_no += 1

    print(f"  Flag column: {n_yes} YES, {n_no} NO")

    # Conditional formatting: highlight NO rows (entire row) with light orange fill
    # We apply a formula rule to the full data range that checks if column V = "NO"
    max_row = max(corrections.keys())
    data_range = f"A3:{get_column_letter(w_col)}{max_row}"
    formula = f'${v_letter}3="NO"'

    rule = FormulaRule(
        formula=[formula],
        fill=ORANGE,
    )
    ws_h.conditional_formatting.add(data_range, rule)
    print(f"  Conditional format applied to {data_range}")

    return v_col, w_col


def update_banner(ws_h):
    """Update Human_Data_Entry A1 banner."""
    cell = ws_h.cell(1, 1)
    cell.value = CORRECTED_BANNER
    cell.font = BOLD
    cell.fill = YELLOW
    print("  Banner updated")


def update_corrections_log(ws_cl):
    """Append a row to Corrections_Log summarising the 294/126 split."""
    # Find the last non-empty row
    last_row = 1
    for r in range(1, 200):
        if ws_cl.cell(r, 1).value is not None:
            last_row = r

    target = last_row + 2
    cell = ws_cl.cell(target, 1)
    cell.value = (
        f"Price Basis Corrected flag (added 2026-08-13): {N_CORRECTED} of {N_TOTAL} human rows "
        f"re-anchored to release_date from EDGAR. {N_NOT} rows remain on original report_date "
        f"basis (human-only events not in LLM universe, EDGAR lookup deferred). Column "
        f"'Price Basis Corrected' in Human_Data_Entry shows YES/NO per row. NO rows are "
        f"highlighted in light orange. Source: workbook_human_prices_corrected.csv."
    )
    cell.font = ITALIC
    print(f"  Corrections_Log updated at row {target}")


def col(n): return get_column_letter(n)

# Column refs for Accuracy_Conventions formulas (reused from fix_workbooks_v2)
LR = "LLM_Data_Entry"
HR = "Human_Data_Entry"

LLM_F = f"{LR}!$F$3:$F$430"
LLM_H = f"{LR}!$H$3:$H$430"
LLM_Q = f"{LR}!$Q$3:$Q$430"
LLM_R = f"{LR}!$R$3:$R$430"

HUM_G = f"{HR}!$G$3:$G$524"
HUM_J = f"{HR}!$J$3:$J$524"
HUM_R = f"{HR}!$R$3:$R$524"
HUM_S = f"{HR}!$S$3:$S$524"


def pval_formula(yes_col, n_col, floor_col, r):
    return (f'=IFERROR(1-BINOM.DIST({col(yes_col)}{r}-1,'
            f'{col(n_col)}{r},{col(floor_col)}{r},TRUE),"")')

def mde_formula(floor_col, n_col, r):
    return (f'=IFERROR(2*1.645*SQRT({col(floor_col)}{r}'
            f'*(1-{col(floor_col)}{r})/{col(n_col)}{r}),"")')


def add_294_section(ws_ac, v_col_letter):
    """
    Add a new section to Accuracy_Conventions showing human figures for the
    294-corrected rows only, using Human_Data_Entry column V as an extra filter.
    """
    # Find last used row
    last_row = 1
    for r in range(1, 200):
        has_content = False
        for c in range(1, 10):
            if ws_ac.cell(r, c).value is not None:
                has_content = True
                break
        if has_content:
            last_row = r

    start = last_row + 3

    # Reference to the flag column in Human_Data_Entry
    HUM_V = f"{HR}!${v_col_letter}$3:${v_col_letter}$524"

    # Title
    title_cell = ws_ac.cell(start, 1)
    title_cell.value = (
        f"Human arm — 294 corrected rows only (Price Basis Corrected = YES)"
    )
    title_cell.font = BOLD
    title_cell.fill = GREY

    note = ws_ac.cell(start + 1, 1)
    note.value = (
        "These figures re-run the coverage and selectivity tables above, restricted to the 294 "
        "human rows where prices have been re-anchored to release_date. The 126 rows still on "
        "original report_date basis are excluded. LLM figures are unchanged (single basis). "
        "The mixed-basis effect on human accuracy is visible by comparing these rows to the "
        "all-420-rows figures above."
    )
    note.font = ITALIC

    r = start + 3

    # ── Coverage table (294 only) ──────────────────────────────────────────────
    ws_ac.cell(r, 1).value = "Coverage (294-corrected rows only)"
    ws_ac.cell(r, 1).font = BOLD
    r += 1
    ws_ac.cell(r, 1).value = "Arm"
    ws_ac.cell(r, 2).value = "YES"
    ws_ac.cell(r, 3).value = "N"
    ws_ac.cell(r, 4).value = "Accuracy"
    ws_ac.cell(r, 5).value = "Floor (always-DOWN)"
    ws_ac.cell(r, 6).value = "p-value"
    ws_ac.cell(r, 7).value = "MDE"
    for c in range(1, 8):
        ws_ac.cell(r, c).font = BOLD
    r += 1

    # Human all (294-corrected)
    n_col_h  = 3
    yes_col_h = 2
    floor_col_h = 5
    ws_ac.cell(r, 1).value = "Human (all raters, 294 corrected)"
    ws_ac.cell(r, yes_col_h).value = (
        f'=COUNTIFS({HUM_S},"YES",{HUM_V},"YES")'
    )
    ws_ac.cell(r, n_col_h).value = (
        f'=COUNTIFS({HUM_S},"<>\"\"",{HUM_V},"YES")'
    )
    ws_ac.cell(r, 4).value = (
        f'=IFERROR({col(yes_col_h)}{r}/{col(n_col_h)}{r},"")'
    )
    ws_ac.cell(r, floor_col_h).value = (
        f'=IFERROR(COUNTIFS({HUM_R},"DOWN",{HUM_V},"YES")'
        f'/COUNTIFS({HUM_S},"<>\"\"",{HUM_V},"YES"),"")'
    )
    ws_ac.cell(r, 6).value = pval_formula(yes_col_h, n_col_h, floor_col_h, r)
    ws_ac.cell(r, 7).value = mde_formula(floor_col_h, n_col_h, r)
    r += 1

    # LLM (unchanged — note only)
    ws_ac.cell(r, 1).value = "LLM (single basis, unchanged)"
    ws_ac.cell(r, 2).value = f'=COUNTIFS({LLM_F},"LLM",{LLM_R},"YES")'
    ws_ac.cell(r, 3).value = f'=COUNTIFS({LLM_F},"LLM",{LLM_R},"<>""")'
    ws_ac.cell(r, 4).value = f'=IFERROR({col(2)}{r}/{col(3)}{r},"")'
    ws_ac.cell(r, 5).value = f'=IFERROR(COUNTIFS({LLM_F},"LLM",{LLM_Q},"DOWN")/COUNTIFS({LLM_F},"LLM",{LLM_R},"<>"""),"")'
    ws_ac.cell(r, 6).value = pval_formula(2, 3, 5, r)
    ws_ac.cell(r, 7).value = mde_formula(5, 3, r)
    r += 2

    # ── Selectivity table (294 only) ──────────────────────────────────────────
    ws_ac.cell(r, 1).value = "Selectivity (294-corrected rows only)"
    ws_ac.cell(r, 1).font = BOLD
    r += 1
    ws_ac.cell(r, 1).value = "Arm"
    ws_ac.cell(r, 2).value = "Correct"
    ws_ac.cell(r, 3).value = "N (traded+graded)"
    ws_ac.cell(r, 4).value = "Accuracy"
    ws_ac.cell(r, 5).value = "Floor (always-DOWN)"
    ws_ac.cell(r, 6).value = "p-value"
    ws_ac.cell(r, 7).value = "MDE"
    for c in range(1, 8):
        ws_ac.cell(r, c).font = BOLD
    r += 1

    # Human selectivity (294 only)
    # Selectivity = BUY/SELL AND UP/DOWN AND flag=YES
    ws_ac.cell(r, 1).value = "Human (all raters, 294 corrected)"
    sel_n = (
        f'=COUNTIFS({HUM_J},"BUY",{HUM_R},"UP",{HUM_V},"YES")'
        f'+COUNTIFS({HUM_J},"BUY",{HUM_R},"DOWN",{HUM_V},"YES")'
        f'+COUNTIFS({HUM_J},"SELL",{HUM_R},"UP",{HUM_V},"YES")'
        f'+COUNTIFS({HUM_J},"SELL",{HUM_R},"DOWN",{HUM_V},"YES")'
    )
    sel_yes = (
        f'=COUNTIFS({HUM_J},"BUY",{HUM_R},"UP",{HUM_V},"YES")'
        f'+COUNTIFS({HUM_J},"SELL",{HUM_R},"DOWN",{HUM_V},"YES")'
    )
    sel_floor = (
        f'=IFERROR((COUNTIFS({HUM_J},"BUY",{HUM_R},"DOWN",{HUM_V},"YES")'
        f'+COUNTIFS({HUM_J},"SELL",{HUM_R},"DOWN",{HUM_V},"YES"))'
        f'/({sel_n.lstrip("=")}),"")'
    )
    ws_ac.cell(r, 2).value = sel_yes
    ws_ac.cell(r, 3).value = sel_n
    ws_ac.cell(r, 4).value = f'=IFERROR({col(2)}{r}/{col(3)}{r},"")'
    ws_ac.cell(r, 5).value = sel_floor
    ws_ac.cell(r, 6).value = pval_formula(2, 3, 5, r)
    ws_ac.cell(r, 7).value = mde_formula(5, 3, r)
    r += 1

    # LLM selectivity (unchanged)
    ws_ac.cell(r, 1).value = "LLM (single basis, unchanged)"
    llm_sel_n = (
        f'=COUNTIFS({LLM_F},"LLM",{LLM_H},"BUY",{LLM_Q},"UP")'
        f'+COUNTIFS({LLM_F},"LLM",{LLM_H},"BUY",{LLM_Q},"DOWN")'
        f'+COUNTIFS({LLM_F},"LLM",{LLM_H},"SELL",{LLM_Q},"UP")'
        f'+COUNTIFS({LLM_F},"LLM",{LLM_H},"SELL",{LLM_Q},"DOWN")'
    )
    llm_sel_yes = (
        f'=COUNTIFS({LLM_F},"LLM",{LLM_H},"BUY",{LLM_Q},"UP")'
        f'+COUNTIFS({LLM_F},"LLM",{LLM_H},"SELL",{LLM_Q},"DOWN")'
    )
    ws_ac.cell(r, 2).value = llm_sel_yes
    ws_ac.cell(r, 3).value = llm_sel_n
    ws_ac.cell(r, 4).value = f'=IFERROR({col(2)}{r}/{col(3)}{r},"")'
    ws_ac.cell(r, 5).value = (
        f'=IFERROR((COUNTIFS({LLM_F},"LLM",{LLM_H},"BUY",{LLM_Q},"DOWN")'
        f'+COUNTIFS({LLM_F},"LLM",{LLM_H},"SELL",{LLM_Q},"DOWN"))'
        f'/({llm_sel_n.lstrip("=")}),"")'
    )
    ws_ac.cell(r, 6).value = pval_formula(2, 3, 5, r)
    ws_ac.cell(r, 7).value = mde_formula(5, 3, r)
    r += 2

    note2 = ws_ac.cell(r, 1)
    note2.value = (
        "Interpretation: If the 294-corrected figures differ materially from the all-420 figures "
        "above, the 126 uncorrected rows are shifting the headline. Report both when citing human "
        "arm accuracy from this workbook."
    )
    note2.font = ITALIC
    note2.fill = YELLOW

    print(f"  Accuracy_Conventions: 294-corrected section added at row {start}")


def main():
    print(f"Loading corrections from {HUMAN_PRICES.name} ...")
    corrections = load_price_corrections()
    print(f"  {sum(1 for v in corrections.values() if v['correctable']=='TRUE')} TRUE, "
          f"{sum(1 for v in corrections.values() if v['correctable']=='FALSE')} FALSE")

    print(f"\nOpening {OUT_CORRECTED.name} ...")
    wb = openpyxl.load_workbook(OUT_CORRECTED)

    ws_h  = wb["Human_Data_Entry"]
    ws_cl = wb["Corrections_Log"]
    ws_ac = wb["Accuracy_Conventions"]

    # 1. Add flag columns, get column index for V
    print("\n1. Adding flag columns to Human_Data_Entry ...")
    v_col, w_col = add_flag_columns(ws_h, corrections)
    v_letter = get_column_letter(v_col)

    # 2. Update banner
    print("\n2. Updating banner ...")
    update_banner(ws_h)

    # 3. Update Corrections_Log
    print("\n3. Updating Corrections_Log ...")
    update_corrections_log(ws_cl)

    # 4. Update Accuracy_Conventions with 294-only section
    print("\n4. Updating Accuracy_Conventions ...")
    add_294_section(ws_ac, v_letter)

    # Save
    print(f"\nSaving {OUT_CORRECTED.name} ...")
    wb.save(OUT_CORRECTED)
    print("Done.")

    # Verify flag column is correct
    wb2 = openpyxl.load_workbook(OUT_CORRECTED, data_only=True)
    ws_check = wb2["Human_Data_Entry"]
    n_yes = sum(1 for ri in corrections if ws_check.cell(ri, v_col).value == "YES")
    n_no  = sum(1 for ri in corrections if ws_check.cell(ri, v_col).value == "NO")
    print(f"\nPost-save verification: {n_yes} YES, {n_no} NO in column {v_letter}")
    if n_yes != N_CORRECTED or n_no != N_NOT:
        print(f"  WARNING: expected {N_CORRECTED} YES, {N_NOT} NO")
    else:
        print("  Counts match. Flag column verified.")


if __name__ == "__main__":
    main()
