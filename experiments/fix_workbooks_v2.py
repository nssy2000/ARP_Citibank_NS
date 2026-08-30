"""
Fix workbooks after post-build review (2026-08-13).

Items addressed:
1. Fix CORRECTED banner in Human_Data_Entry A1 (wrong banner was used)
2. Build Accuracy_Conventions sheet in both workbooks (live formulas)
3. Add plain-text note to Charts above Figure 1 (denominator mismatch warning)
4. Revert LLM_Data_Entry in LOCKED workbook to original source values
5. Add LLM alignment assertion (previously unreported)
6. Report verification

Selectivity cross-check finding (reported before proceeding):
  Live formula gives 61/94 = 64.9%; repo authoritative = 62/95 = 65.3%.
  Diff: 1 event. Cause: one event classified FLAT in current returns_matrix
  but apparently counted as graded in ext2 run (20260812_183842). No event
  has |ret| between 2.00% and 2.01%; closest is Puma|2025|Q3 at 1.9903%.
  The live formula definition is correct (H!=HOLD, Q=UP/DOWN); the 1-event
  deviation is documented in a note on the sheet.
"""

import csv, datetime, json, shutil, sys
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference

ROOT   = Path(__file__).resolve().parent.parent
WB_DIR = ROOT / "data" / "workbook"
SUMMARY = ROOT / "outputs" / "global" / "summary"

SOURCE_WB    = WB_DIR / "Master_Data_NEW_REPAIRED_2026-08-09.xlsx"
OUT_LOCKED   = WB_DIR / "Master_Data_LOCKED_2026-08-13.xlsx"
OUT_CORRECTED = WB_DIR / "Master_Data_CORRECTED_2026-08-13.xlsx"

BOLD   = Font(bold=True)
ITALIC = Font(italic=True)
GREY   = PatternFill("solid", fgColor="D9D9D9")
YELLOW = PatternFill("solid", fgColor="FFFFC0")
RED    = PatternFill("solid", fgColor="FFCCCC")

CORRECTED_BANNER = (
    "HUMAN ARM CORRECTED. Prices M to P re-anchored on release_date from EDGAR filing dates, "
    "2026-08-13. Columns Q to U and the Summary, Charts and Efficiency sheets recalculate from "
    "these. Original basis retained in Master_Data_LOCKED_2026-08-13.xlsx. See Corrections_Log."
)

LOCKED_BANNER = (
    "HUMAN ARM FROZEN. Prices and derived accuracy on the original report_date basis, "
    "superseded 2026-08-12. Net P&L in column U is a formula deriving from those prices "
    "and therefore remains on the superseded basis by design. Corrected figures are in "
    "Corrections_Log and Corrected_Metrics."
)

# Column refs (1-based)
# LLM_Data_Entry
L_F = 6   # Type (LLM)
L_H = 8   # Decision
L_Q = 17  # Actual Direction (UP/DOWN/FLAT)
L_R = 18  # Prediction Correct? (YES/NO/FLAT-equivalent)

# Human_Data_Entry
H_G = 7   # Rater name
H_J = 10  # Decision (formula from sentiment)
H_R = 18  # Actual Direction (formula)
H_S = 19  # Prediction Correct? (formula)


# ─────────────────────────────────────────────────────────────────────────────
# LLM alignment assertion
# ─────────────────────────────────────────────────────────────────────────────

def assert_llm_alignment(ws_l, llm_rows):
    """
    Verify that for each LLM row written, the Company name in the sheet
    matches what the llm_rows record expects.
    Returns list of failure strings.
    """
    failures = []
    for rec in llm_rows:
        r = rec["row_index"]
        sheet_co = ws_l.cell(r, 1).value   # Column A = Company
        expect_co = rec["company"]
        if sheet_co != expect_co:
            failures.append(
                f"Row {r}: sheet Company={sheet_co!r}, expected={expect_co!r}"
            )
    return failures


# ─────────────────────────────────────────────────────────────────────────────
# Charts sheet: add denominator-mismatch note above Figure 1
# ─────────────────────────────────────────────────────────────────────────────

NOTE_TEXT = (
    "DENOMINATOR NOTE (added 2026-08-13): the bars in Figure 1 compute coverage accuracy "
    "(YES / all events including HOLDs). The floor in row 14 is the majority-direction count "
    "over human-arm events only, not LLM events. These denominators are not matched. For "
    "matched coverage vs selectivity accuracy with each floor computed on its own "
    "denominator, see the Accuracy_Conventions sheet."
)

def add_charts_note(ws_c):
    """Write a plain-text note into Charts row 3 (currently blank)."""
    cell = ws_c.cell(3, 1)
    cell.value = NOTE_TEXT
    cell.font = ITALIC
    cell.fill = YELLOW


# ─────────────────────────────────────────────────────────────────────────────
# Accuracy_Conventions sheet
# ─────────────────────────────────────────────────────────────────────────────

# Excel column letters for formula construction (1-based index → letter)
def col(n): return get_column_letter(n)

# LLM_Data_Entry ranges
LR = "LLM_Data_Entry"
LLM_F  = f"{LR}!$F$3:$F$430"
LLM_H  = f"{LR}!$H$3:$H$430"
LLM_Q  = f"{LR}!$Q$3:$Q$430"
LLM_R  = f"{LR}!$R$3:$R$430"

# Human_Data_Entry ranges
HR = "Human_Data_Entry"
HUM_G  = f"{HR}!$G$3:$G$524"
HUM_J  = f"{HR}!$J$3:$J$524"
HUM_R  = f"{HR}!$R$3:$R$524"
HUM_S  = f"{HR}!$S$3:$S$524"


def llm_cov_n_formula():
    return f'=COUNTIFS({LLM_F},"LLM",{LLM_R},"<>""")'

def llm_cov_yes_formula():
    return f'=COUNTIFS({LLM_F},"LLM",{LLM_R},"YES")'

def llm_cov_floor_down_formula():
    return f'=COUNTIFS({LLM_F},"LLM",{LLM_Q},"DOWN")'

def llm_hold_n_formula():
    return f'=COUNTIFS({LLM_F},"LLM",{LLM_H},"HOLD")'

def llm_traded_formula():
    return (f'=COUNTIFS({LLM_F},"LLM",{LLM_H},"BUY")'
            f'+COUNTIFS({LLM_F},"LLM",{LLM_H},"SELL")')

def llm_sel_n_formula():
    # Traded (BUY/SELL) AND graded (UP/DOWN)
    return (f'=COUNTIFS({LLM_F},"LLM",{LLM_H},"BUY",{LLM_Q},"UP")'
            f'+COUNTIFS({LLM_F},"LLM",{LLM_H},"BUY",{LLM_Q},"DOWN")'
            f'+COUNTIFS({LLM_F},"LLM",{LLM_H},"SELL",{LLM_Q},"UP")'
            f'+COUNTIFS({LLM_F},"LLM",{LLM_H},"SELL",{LLM_Q},"DOWN")')

def llm_sel_yes_formula():
    return (f'=COUNTIFS({LLM_F},"LLM",{LLM_H},"BUY",{LLM_Q},"UP")'
            f'+COUNTIFS({LLM_F},"LLM",{LLM_H},"SELL",{LLM_Q},"DOWN")')

def llm_sel_floor_down_formula():
    # DOWN outcomes among traded graded
    return (f'=COUNTIFS({LLM_F},"LLM",{LLM_H},"BUY",{LLM_Q},"DOWN")'
            f'+COUNTIFS({LLM_F},"LLM",{LLM_H},"SELL",{LLM_Q},"DOWN")')

def human_all_cov_n():
    return f'=COUNTIF({HUM_S},"<>""")'

def human_all_cov_yes():
    return f'=COUNTIF({HUM_S},"YES")'

def human_all_cov_floor_down():
    return f'=COUNTIF({HUM_R},"DOWN")'

def human_all_hold_n():
    return f'=COUNTIF({HUM_J},"HOLD")'

def human_all_traded():
    return f'=COUNTIF({HUM_J},"BUY")+COUNTIF({HUM_J},"SELL")'

def human_all_sel_n():
    return (f'=COUNTIFS({HUM_J},"BUY",{HUM_R},"UP")'
            f'+COUNTIFS({HUM_J},"BUY",{HUM_R},"DOWN")'
            f'+COUNTIFS({HUM_J},"SELL",{HUM_R},"UP")'
            f'+COUNTIFS({HUM_J},"SELL",{HUM_R},"DOWN")')

def human_all_sel_yes():
    return (f'=COUNTIFS({HUM_J},"BUY",{HUM_R},"UP")'
            f'+COUNTIFS({HUM_J},"SELL",{HUM_R},"DOWN")')

def human_all_sel_floor_down():
    return (f'=COUNTIFS({HUM_J},"BUY",{HUM_R},"DOWN")'
            f'+COUNTIFS({HUM_J},"SELL",{HUM_R},"DOWN")')

def rater_cov_n(a_col, a_row):
    ref = f"{col(a_col)}{a_row}"
    return f'=COUNTIFS({HUM_G},{ref},{HUM_S},"<>""")'

def rater_cov_yes(a_col, a_row):
    ref = f"{col(a_col)}{a_row}"
    return f'=COUNTIFS({HUM_G},{ref},{HUM_S},"YES")'

def rater_cov_floor_down(a_col, a_row):
    ref = f"{col(a_col)}{a_row}"
    return f'=COUNTIFS({HUM_G},{ref},{HUM_R},"DOWN")'

def rater_hold_n(a_col, a_row):
    ref = f"{col(a_col)}{a_row}"
    return f'=COUNTIFS({HUM_G},{ref},{HUM_J},"HOLD")'

def rater_traded(a_col, a_row):
    ref = f"{col(a_col)}{a_row}"
    return (f'=COUNTIFS({HUM_G},{ref},{HUM_J},"BUY")'
            f'+COUNTIFS({HUM_G},{ref},{HUM_J},"SELL")')

def rater_sel_n(a_col, a_row):
    ref = f"{col(a_col)}{a_row}"
    return (f'=COUNTIFS({HUM_G},{ref},{HUM_J},"BUY",{HUM_R},"UP")'
            f'+COUNTIFS({HUM_G},{ref},{HUM_J},"BUY",{HUM_R},"DOWN")'
            f'+COUNTIFS({HUM_G},{ref},{HUM_J},"SELL",{HUM_R},"UP")'
            f'+COUNTIFS({HUM_G},{ref},{HUM_J},"SELL",{HUM_R},"DOWN")')

def rater_sel_yes(a_col, a_row):
    ref = f"{col(a_col)}{a_row}"
    return (f'=COUNTIFS({HUM_G},{ref},{HUM_J},"BUY",{HUM_R},"UP")'
            f'+COUNTIFS({HUM_G},{ref},{HUM_J},"SELL",{HUM_R},"DOWN")')

def rater_sel_floor_down(a_col, a_row):
    ref = f"{col(a_col)}{a_row}"
    return (f'=COUNTIFS({HUM_G},{ref},{HUM_J},"BUY",{HUM_R},"DOWN")'
            f'+COUNTIFS({HUM_G},{ref},{HUM_J},"SELL",{HUM_R},"DOWN")')

def acc_formula(yes_col, n_col, r):
    return f'=IF({col(n_col)}{r}=0,"",{col(yes_col)}{r}/{col(n_col)}{r})'

def floor_prob_formula(down_col, n_col, r):
    return f'=IF({col(n_col)}{r}=0,"",{col(down_col)}{r}/{col(n_col)}{r})'

def pval_formula(yes_col, n_col, floor_col, r):
    return (f'=IFERROR(1-BINOM.DIST({col(yes_col)}{r}-1,'
            f'{col(n_col)}{r},{col(floor_col)}{r},TRUE),"")')

def mde_formula(floor_col, n_col, r):
    return (f'=IFERROR(2*1.645*SQRT({col(floor_col)}{r}'
            f'*(1-{col(floor_col)}{r})/{col(n_col)}{r}),"")')

def hold_rate_formula(hold_col, traded_col, r):
    return (f'=IFERROR({col(hold_col)}{r}'
            f'/({col(hold_col)}{r}+{col(traded_col)}{r}),"")')


# Layout columns (1-based):
#  A=1  Arm
#  B=2  n (denominator)
#  C=3  YES (numerator)
#  D=4  Accuracy
#  E=5  floor n_DOWN
#  F=6  Floor prob
#  G=7  p-value
#  H=8  MDE
#  I=9  HOLD count
#  J=10 HOLD rate
#  K=11 Traded count
AC_ARM   = 1
AC_N     = 2
AC_YES   = 3
AC_ACC   = 4
AC_FDOWN = 5
AC_FPROB = 6
AC_PVAL  = 7
AC_MDE   = 8
AC_HOLD  = 9
AC_HRATE = 10
AC_TRAD  = 11

HEADERS = ["Arm", "n (denom)", "YES", "Accuracy", "Floor n_DOWN", "Floor prob",
           "p-value (1-sided)", "MDE (±, 90%)", "HOLD count", "HOLD rate", "Traded"]

RATER_REFS = [  # Settings cell references for 7 human raters
    "=Settings!$A$8", "=Settings!$A$9", "=Settings!$A$10",
    "=Settings!$A$11", "=Settings!$A$12", "=Settings!$A$13", "=Settings!$A$14",
]


def build_accuracy_conventions(wb):
    ws = wb.create_sheet(title="Accuracy_Conventions")

    # Column widths
    ws.column_dimensions["A"].width = 20
    for c in "BCDEFGHIJK":
        ws.column_dimensions[c].width = 14

    def h(r, c, v, bold=False, fill=None, italic=False):
        cell = ws.cell(r, c)
        cell.value = v
        if bold: cell.font = Font(bold=True)
        elif italic: cell.font = Font(italic=True)
        if fill: cell.fill = fill

    # ── Row 1: Title ──────────────────────────────────────────────────────────
    h(1, 1, "Accuracy Conventions — Coverage vs Selectivity", bold=True)

    # ── Row 2: Plain-text note (item 2 requirement) ───────────────────────────
    h(2, 1, (
        "Coverage accuracy = YES / all events (HOLDs counted as failures when stock moved, "
        "credited as correct when stock stayed flat). "
        "Selectivity accuracy = correct direction / (traded-and-graded events only: "
        "BUY or SELL call AND |ret_overnight| > 2%). "
        "Neither figure travels alone: selectivity inflates when the HOLD rate is high, "
        "because the model can selectively avoid hard calls. "
        "The FLAT convention gives the LLM a ~7.1pp advantage in the paired comparison "
        "because LLM holds 40.4% of events vs human arm 23.4% (N=171 paired subset): "
        "when the LLM holds and the stock stays flat, it earns a YES; a human BUY/SELL "
        "on the same event would earn a NO. "
        "See methodology_flat_convention.md for full derivation."
    ), italic=True)

    # ── Row 3: Cross-check note ───────────────────────────────────────────────
    h(3, 1, (
        "CROSS-CHECK NOTE: the selectivity formula (H=BUY/SELL, Q=UP/DOWN) gives "
        "61/94 = 64.9% for the LLM arm. The repo authoritative figure is 62/95 = 65.3% "
        "(from item_e_walkforward.json, run 20260812_183842). Discrepancy = 1 event. "
        "Cause: one event's |ret_overnight| is below the ±2% band in the current "
        "returns_matrix but was apparently above it in the ext2 run. "
        "The formula definition is correct; the 0.4pp deviation is documented here."
    ), italic=True, fill=YELLOW)

    # ── Row 5: Coverage table ─────────────────────────────────────────────────
    h(5, 1, "COVERAGE ACCURACY (YES / all events)", bold=True, fill=GREY)
    ws.merge_cells(start_row=5, start_column=1, end_row=5, end_column=11)

    h(6, 1, "Denominator: all events for the arm (including HOLDs)", italic=True)
    h(6, 6, "Floor: majority direction / all events", italic=True)

    # Header row 7
    for ci, hdr in enumerate(HEADERS, 1):
        h(7, ci, hdr, bold=True, fill=GREY)

    # LLM row 8
    ROW_COV_LLM = 8
    ws.cell(ROW_COV_LLM, AC_ARM).value = "LLM (DeepSeek)"
    ws.cell(ROW_COV_LLM, AC_N).value    = llm_cov_n_formula()
    ws.cell(ROW_COV_LLM, AC_YES).value  = llm_cov_yes_formula()
    ws.cell(ROW_COV_LLM, AC_ACC).value  = acc_formula(AC_YES, AC_N, ROW_COV_LLM)
    ws.cell(ROW_COV_LLM, AC_FDOWN).value= llm_cov_floor_down_formula()
    ws.cell(ROW_COV_LLM, AC_FPROB).value= floor_prob_formula(AC_FDOWN, AC_N, ROW_COV_LLM)
    ws.cell(ROW_COV_LLM, AC_PVAL).value = pval_formula(AC_YES, AC_N, AC_FPROB, ROW_COV_LLM)
    ws.cell(ROW_COV_LLM, AC_MDE).value  = mde_formula(AC_FPROB, AC_N, ROW_COV_LLM)
    ws.cell(ROW_COV_LLM, AC_HOLD).value = llm_hold_n_formula()
    ws.cell(ROW_COV_LLM, AC_TRAD).value = llm_traded_formula()
    ws.cell(ROW_COV_LLM, AC_HRATE).value= hold_rate_formula(AC_HOLD, AC_TRAD, ROW_COV_LLM)

    # Human All row 9
    ROW_COV_HALL = 9
    ws.cell(ROW_COV_HALL, AC_ARM).value = "Human (All)"
    ws.cell(ROW_COV_HALL, AC_N).value    = human_all_cov_n()
    ws.cell(ROW_COV_HALL, AC_YES).value  = human_all_cov_yes()
    ws.cell(ROW_COV_HALL, AC_ACC).value  = acc_formula(AC_YES, AC_N, ROW_COV_HALL)
    ws.cell(ROW_COV_HALL, AC_FDOWN).value= human_all_cov_floor_down()
    ws.cell(ROW_COV_HALL, AC_FPROB).value= floor_prob_formula(AC_FDOWN, AC_N, ROW_COV_HALL)
    ws.cell(ROW_COV_HALL, AC_PVAL).value = pval_formula(AC_YES, AC_N, AC_FPROB, ROW_COV_HALL)
    ws.cell(ROW_COV_HALL, AC_MDE).value  = mde_formula(AC_FPROB, AC_N, ROW_COV_HALL)
    ws.cell(ROW_COV_HALL, AC_HOLD).value = human_all_hold_n()
    ws.cell(ROW_COV_HALL, AC_TRAD).value = human_all_traded()
    ws.cell(ROW_COV_HALL, AC_HRATE).value= hold_rate_formula(AC_HOLD, AC_TRAD, ROW_COV_HALL)

    # Individual rater rows 10-16
    ROW_COV_RATERS = list(range(10, 17))
    for i, (row, rater_ref) in enumerate(zip(ROW_COV_RATERS, RATER_REFS)):
        ws.cell(row, AC_ARM).value = rater_ref
        ws.cell(row, AC_N).value    = rater_cov_n(AC_ARM, row)
        ws.cell(row, AC_YES).value  = rater_cov_yes(AC_ARM, row)
        ws.cell(row, AC_ACC).value  = acc_formula(AC_YES, AC_N, row)
        ws.cell(row, AC_FDOWN).value= rater_cov_floor_down(AC_ARM, row)
        ws.cell(row, AC_FPROB).value= floor_prob_formula(AC_FDOWN, AC_N, row)
        ws.cell(row, AC_PVAL).value = pval_formula(AC_YES, AC_N, AC_FPROB, row)
        ws.cell(row, AC_MDE).value  = mde_formula(AC_FPROB, AC_N, row)
        ws.cell(row, AC_HOLD).value = rater_hold_n(AC_ARM, row)
        ws.cell(row, AC_TRAD).value = rater_traded(AC_ARM, row)
        ws.cell(row, AC_HRATE).value= hold_rate_formula(AC_HOLD, AC_TRAD, row)

    # ── Row 18: Selectivity table ─────────────────────────────────────────────
    h(18, 1, "SELECTIVITY ACCURACY (correct / traded-and-graded events)", bold=True, fill=GREY)
    ws.merge_cells(start_row=18, start_column=1, end_row=18, end_column=11)

    h(19, 1, "Denominator: events where arm committed (BUY/SELL) AND |ret_overnight| > 2%", italic=True)
    h(19, 6, "Floor: DOWN outcomes / graded denominator", italic=True)

    # Header row 20
    for ci, hdr in enumerate(HEADERS, 1):
        h(20, ci, hdr, bold=True, fill=GREY)

    # LLM row 21
    ROW_SEL_LLM = 21
    ws.cell(ROW_SEL_LLM, AC_ARM).value = "LLM (DeepSeek)"
    ws.cell(ROW_SEL_LLM, AC_N).value    = llm_sel_n_formula()
    ws.cell(ROW_SEL_LLM, AC_YES).value  = llm_sel_yes_formula()
    ws.cell(ROW_SEL_LLM, AC_ACC).value  = acc_formula(AC_YES, AC_N, ROW_SEL_LLM)
    ws.cell(ROW_SEL_LLM, AC_FDOWN).value= llm_sel_floor_down_formula()
    ws.cell(ROW_SEL_LLM, AC_FPROB).value= floor_prob_formula(AC_FDOWN, AC_N, ROW_SEL_LLM)
    ws.cell(ROW_SEL_LLM, AC_PVAL).value = pval_formula(AC_YES, AC_N, AC_FPROB, ROW_SEL_LLM)
    ws.cell(ROW_SEL_LLM, AC_MDE).value  = mde_formula(AC_FPROB, AC_N, ROW_SEL_LLM)
    ws.cell(ROW_SEL_LLM, AC_HOLD).value = llm_hold_n_formula()
    ws.cell(ROW_SEL_LLM, AC_TRAD).value = llm_traded_formula()
    ws.cell(ROW_SEL_LLM, AC_HRATE).value= hold_rate_formula(AC_HOLD, AC_TRAD, ROW_SEL_LLM)

    # Human All row 22
    ROW_SEL_HALL = 22
    ws.cell(ROW_SEL_HALL, AC_ARM).value = "Human (All)"
    ws.cell(ROW_SEL_HALL, AC_N).value    = human_all_sel_n()
    ws.cell(ROW_SEL_HALL, AC_YES).value  = human_all_sel_yes()
    ws.cell(ROW_SEL_HALL, AC_ACC).value  = acc_formula(AC_YES, AC_N, ROW_SEL_HALL)
    ws.cell(ROW_SEL_HALL, AC_FDOWN).value= human_all_sel_floor_down()
    ws.cell(ROW_SEL_HALL, AC_FPROB).value= floor_prob_formula(AC_FDOWN, AC_N, ROW_SEL_HALL)
    ws.cell(ROW_SEL_HALL, AC_PVAL).value = pval_formula(AC_YES, AC_N, AC_FPROB, ROW_SEL_HALL)
    ws.cell(ROW_SEL_HALL, AC_MDE).value  = mde_formula(AC_FPROB, AC_N, ROW_SEL_HALL)
    ws.cell(ROW_SEL_HALL, AC_HOLD).value = human_all_hold_n()
    ws.cell(ROW_SEL_HALL, AC_TRAD).value = human_all_traded()
    ws.cell(ROW_SEL_HALL, AC_HRATE).value= hold_rate_formula(AC_HOLD, AC_TRAD, ROW_SEL_HALL)

    # Individual rater rows 23-29
    ROW_SEL_RATERS = list(range(23, 30))
    for i, (row, rater_ref) in enumerate(zip(ROW_SEL_RATERS, RATER_REFS)):
        ws.cell(row, AC_ARM).value = rater_ref
        ws.cell(row, AC_N).value    = rater_sel_n(AC_ARM, row)
        ws.cell(row, AC_YES).value  = rater_sel_yes(AC_ARM, row)
        ws.cell(row, AC_ACC).value  = acc_formula(AC_YES, AC_N, row)
        ws.cell(row, AC_FDOWN).value= rater_sel_floor_down(AC_ARM, row)
        ws.cell(row, AC_FPROB).value= floor_prob_formula(AC_FDOWN, AC_N, row)
        ws.cell(row, AC_PVAL).value = pval_formula(AC_YES, AC_N, AC_FPROB, row)
        ws.cell(row, AC_MDE).value  = mde_formula(AC_FPROB, AC_N, row)
        ws.cell(row, AC_HOLD).value = rater_hold_n(AC_ARM, row)
        ws.cell(row, AC_TRAD).value = rater_traded(AC_ARM, row)
        ws.cell(row, AC_HRATE).value= hold_rate_formula(AC_HOLD, AC_TRAD, row)

    # ── Row 31: Chart data summary ─────────────────────────────────────────────
    h(31, 1, "Chart data (feeds accuracy comparison chart below)", bold=True, fill=GREY)
    ws.merge_cells(start_row=31, start_column=1, end_row=31, end_column=5)

    # Chart data headers row 32
    chart_hdrs = ["Arm", "Coverage acc", "Selectivity acc", "Coverage floor", "Selectivity floor"]
    for ci, hdr in enumerate(chart_hdrs, 1):
        h(32, ci, hdr, bold=True, fill=GREY)

    # Chart data rows 33-34
    # LLM row 33
    ws.cell(33, 1).value = "LLM"
    ws.cell(33, 2).value = f"=D{ROW_COV_LLM}"   # Coverage acc
    ws.cell(33, 3).value = f"=D{ROW_SEL_LLM}"   # Selectivity acc
    ws.cell(33, 4).value = f"=F{ROW_COV_LLM}"   # Coverage floor
    ws.cell(33, 5).value = f"=F{ROW_SEL_LLM}"   # Selectivity floor

    # Human All row 34
    ws.cell(34, 1).value = "Human (All)"
    ws.cell(34, 2).value = f"=D{ROW_COV_HALL}"
    ws.cell(34, 3).value = f"=D{ROW_SEL_HALL}"
    ws.cell(34, 4).value = f"=F{ROW_COV_HALL}"
    ws.cell(34, 5).value = f"=F{ROW_SEL_HALL}"

    # ── Chart ─────────────────────────────────────────────────────────────────
    chart = BarChart()
    chart.type = "col"
    chart.grouping = "clustered"
    chart.title = "Coverage vs Selectivity Accuracy by Arm"
    chart.y_axis.title = "Accuracy"
    chart.x_axis.title = "Arm"
    chart.width = 20
    chart.height = 12

    # Four series: cov_acc (col B), sel_acc (col C), cov_floor (col D), sel_floor (col E)
    # Each from rows 32-34 (header + 2 data rows)
    for series_col, series_name in [
        (2, "Coverage acc"),
        (3, "Selectivity acc"),
        (4, "Coverage floor"),
        (5, "Selectivity floor"),
    ]:
        data = Reference(ws, min_col=series_col, max_col=series_col,
                         min_row=32, max_row=34)
        chart.add_data(data, titles_from_data=True)

    cats = Reference(ws, min_col=1, max_col=1, min_row=33, max_row=34)
    chart.set_categories(cats)

    ws.add_chart(chart, "A36")

    # ── Row 60: Key distinction note ──────────────────────────────────────────
    h(60, 1, (
        "KEY DISTINCTION: Coverage (42%) and selectivity (65%) are not both correct — "
        "they measure different things on different denominators. Coverage penalises every "
        "HOLD call; selectivity credits only committed events that moved. Quoting one "
        "without the other is misleading. The repo's headline accuracy figure is selectivity "
        "(62/95 = 65.3%, from item_e_walkforward.json); the workbook Summary and Figure 1 "
        "show coverage (42% for LLM). Both are real; neither is the whole story."
    ), italic=True)

    return ws


# ─────────────────────────────────────────────────────────────────────────────
# Revert LLM_Data_Entry in locked workbook to original source values
# ─────────────────────────────────────────────────────────────────────────────

def revert_llm_to_source(wb_locked, wb_source):
    """
    Copy all cell values from source LLM_Data_Entry into locked workbook
    LLM_Data_Entry, restoring the pre-correction basis.
    """
    ws_src = wb_source["LLM_Data_Entry"]
    ws_dst = wb_locked["LLM_Data_Entry"]

    n_cells = 0
    max_row = ws_src.max_row
    max_col = ws_src.max_column
    for r in range(1, max_row + 1):
        for c in range(1, max_col + 1):
            src_val = ws_src.cell(r, c).value
            dst_cell = ws_dst.cell(r, c)
            if dst_cell.value != src_val:
                dst_cell.value = src_val
                n_cells += 1
    return n_cells


# ─────────────────────────────────────────────────────────────────────────────
# Main
# ─────────────────────────────────────────────────────────────────────────────

def main():
    print("Loading workbooks...")

    wb_src = openpyxl.load_workbook(SOURCE_WB, data_only=False)
    wb_locked = openpyxl.load_workbook(OUT_LOCKED, data_only=False)
    wb_corrected = openpyxl.load_workbook(OUT_CORRECTED, data_only=False)

    # Load LLM rows from CSV for alignment check
    llm_rows = []
    with open(SUMMARY / "workbook_llm_corrected.csv") as f:
        rd = csv.DictReader(f)
        for row in rd:
            row["row_index"] = int(row["row_index"])
            llm_rows.append(row)
    print(f"  LLM rows loaded: {len(llm_rows)}")

    # ── Item 1: Fix CORRECTED banner ─────────────────────────────────────────
    print("\n=== ITEM 1: Fix CORRECTED banner ===")
    ws_h_c = wb_corrected["Human_Data_Entry"]
    orig_font = ws_h_c.cell(1, 1).font
    ws_h_c.cell(1, 1).value = CORRECTED_BANNER
    ws_h_c.cell(1, 1).font = Font(bold=True,
                                   name=orig_font.name if orig_font else None,
                                   size=orig_font.size if orig_font else None)
    print(f"  CORRECTED A1 → {CORRECTED_BANNER[:60]}...")

    # Verify LOCKED banner is still correct (no change needed)
    ws_h_l = wb_locked["Human_Data_Entry"]
    locked_a1 = ws_h_l.cell(1, 1).value or ""
    if "FROZEN" in locked_a1:
        print("  LOCKED A1 already has FROZEN banner ✓")
    else:
        print(f"  WARNING: LOCKED A1 unexpected: {locked_a1[:60]}")

    # ── LLM alignment assertion (both workbooks) ─────────────────────────────
    print("\n=== LLM ALIGNMENT ASSERTION ===")
    ws_l_c = wb_corrected["LLM_Data_Entry"]
    ws_l_l = wb_locked["LLM_Data_Entry"]

    fails_corrected = assert_llm_alignment(ws_l_c, llm_rows)
    fails_locked    = assert_llm_alignment(ws_l_l, llm_rows)

    print(f"  CORRECTED LLM: {len(fails_corrected)} alignment failures on {len(llm_rows)} rows")
    if fails_corrected:
        for f in fails_corrected[:5]:
            print(f"    {f}")

    print(f"  LOCKED LLM (pre-revert): {len(fails_locked)} alignment failures on {len(llm_rows)} rows")
    if fails_locked:
        for f in fails_locked[:5]:
            print(f"    {f}")

    # ── Item 4: Revert LLM in LOCKED to source ───────────────────────────────
    print("\n=== ITEM 4: Revert LLM_Data_Entry in LOCKED to source ===")
    n_changed = revert_llm_to_source(wb_locked, wb_src)
    print(f"  Cells changed: {n_changed}")

    # Re-run alignment on reverted locked workbook
    ws_l_l2 = wb_locked["LLM_Data_Entry"]
    # After revert, source has Company names in col A, same as what llm_rows expects
    # (The source is what llm_rows was built from)
    fails_locked_post = assert_llm_alignment(ws_l_l2, llm_rows)
    # Note: after revert, locked LLM matches source (pre-correction).
    # llm_rows has company names from returns_matrix which should match source.
    print(f"  LOCKED LLM post-revert alignment: {len(fails_locked_post)} failures")
    if fails_locked_post:
        for f in fails_locked_post[:5]:
            print(f"    {f}")

    # ── Item 3: Add Charts note (both workbooks) ─────────────────────────────
    print("\n=== ITEM 3: Add Charts note ===")
    for wb_name, wb in [("LOCKED", wb_locked), ("CORRECTED", wb_corrected)]:
        ws_c = wb["Charts"]
        add_charts_note(ws_c)
        print(f"  {wb_name}: Charts row 3 note added")

    # ── Item 2: Build Accuracy_Conventions sheet ─────────────────────────────
    print("\n=== ITEM 2: Build Accuracy_Conventions sheet ===")
    for wb_name, wb in [("LOCKED", wb_locked), ("CORRECTED", wb_corrected)]:
        if "Accuracy_Conventions" in wb.sheetnames:
            del wb["Accuracy_Conventions"]
        build_accuracy_conventions(wb)
        print(f"  {wb_name}: Accuracy_Conventions sheet built")

    # ── Save ─────────────────────────────────────────────────────────────────
    print("\n=== Saving workbooks ===")
    wb_locked.save(OUT_LOCKED)
    print(f"  Saved {OUT_LOCKED.name}")
    wb_corrected.save(OUT_CORRECTED)
    print(f"  Saved {OUT_CORRECTED.name}")
    wb_src.close()
    wb_locked.close()
    wb_corrected.close()

    # ── Verification ─────────────────────────────────────────────────────────
    print("\n=== VERIFICATION ===")
    for path in [OUT_LOCKED, OUT_CORRECTED]:
        wb = openpyxl.load_workbook(path, data_only=False)
        sheets = wb.sheetnames
        n_sheets = len(sheets)
        # Expected: 7 original + 7 new + Accuracy_Conventions = 15
        expected = 15

        # Charts count
        n_charts_charts = 0
        n_charts_eff = 0
        n_charts_ac = 0
        if "Charts" in sheets:
            for chart in wb["Charts"]._charts:
                n_charts_charts += 1
        if "Efficiency" in sheets:
            for chart in wb["Efficiency"]._charts:
                n_charts_eff += 1
        if "Accuracy_Conventions" in sheets:
            for chart in wb["Accuracy_Conventions"]._charts:
                n_charts_ac += 1

        # Human S3 formula check
        s3 = wb["Human_Data_Entry"].cell(3, 1).value or ""  # A1 banner in row 1
        h_s3 = wb["Human_Data_Entry"].cell(3, 19).value or ""  # S3 = col 19

        # Banner check
        banner = wb["Human_Data_Entry"].cell(1, 1).value or ""
        banner_ok = "FROZEN" in banner if "LOCKED" in path.name else "CORRECTED" in banner

        print(f"\n  {path.name}:")
        print(f"    Sheet count: {n_sheets} (expected {expected})")
        print(f"    Sheets: {sheets}")
        print(f"    Charts/Charts: {n_charts_charts} (expected 5)")
        print(f"    Charts/Efficiency: {n_charts_eff} (expected 4)")
        print(f"    Charts/Accuracy_Conventions: {n_charts_ac} (expected 1)")
        print(f"    Banner correct: {banner_ok} ({'FROZEN' if 'LOCKED' in path.name else 'CORRECTED'} in A1)")
        print(f"    H_S3 formula: {str(h_s3)[:60]!r}")
        print(f"    H_S3 starts with =IF(OR(N3: {str(h_s3).startswith('=IF(OR(N3')}")

        # Charts note check
        if "Charts" in sheets:
            row3 = wb["Charts"].cell(3, 1).value or ""
            print(f"    Charts row 3 note: {str(row3)[:60]!r}")

        # Accuracy_Conventions cross-check note
        if "Accuracy_Conventions" in sheets:
            r3 = wb["Accuracy_Conventions"].cell(3, 1).value or ""
            print(f"    AC row 3 cross-check note: {str(r3)[:60]!r}")

        # LLM alignment
        ws_l = wb["LLM_Data_Entry"]
        llm_data_rows = sum(1 for r in range(3, ws_l.max_row+1) if ws_l.cell(r,1).value)
        print(f"    LLM_Data_Entry data rows: {llm_data_rows}")

        wb.close()

    print("\n=== IMPORTANT: Circulation note ===")
    print("  Both files contain formula cells that openpyxl stripped cached results from.")
    print("  Prediction Correct? (Human col S, LLM col R), Accuracy_Conventions accuracy,")
    print("  p-values and MDE will all read blank until the file is opened in Excel")
    print("  and Ctrl+Alt+F9 (force full recalculation) is pressed.")
    print("  Do not circulate either file before running recalculation.")

    print("\n=== UNCORRECTABLE ROWS (item 6) ===")
    print("  126 human rows in CORRECTED workbook kept original prices (correctable=FALSE).")
    print("  These rows are NOT visually flagged — they appear identical to corrected rows.")
    print("  A reader cannot distinguish them without consulting workbook_human_prices_corrected.csv.")
    print("  Recommend: the user decides whether to add a flag column or note.")

    print("\nDone.")


if __name__ == "__main__":
    sys.exit(main())
