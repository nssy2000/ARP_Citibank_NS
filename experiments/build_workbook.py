"""
Build four export CSVs and two corrected Excel workbooks.

Outputs (all in outputs/global/summary/ and data/workbook/):
  workbook_human_prices_corrected.csv
  workbook_llm_corrected.csv
  workbook_metrics.csv
  workbook_readme.md
  data/workbook/Master_Data_LOCKED_2026-08-13.xlsx
  data/workbook/Master_Data_CORRECTED_2026-08-13.xlsx

Usage:
  python -m experiments.build_workbook
"""

import csv, datetime, json, math, re, shutil, sys
from pathlib import Path
from collections import defaultdict

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
import pandas as pd

ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT))
from eval.excluded_events import EXCLUDED_EVENTS, EXCLUSION_REASONS  # noqa: E402
SUMMARY = ROOT / "outputs" / "global" / "summary"
WB_DIR  = ROOT / "data" / "workbook"
SOURCE_WB = WB_DIR / "Master_Data_NEW_REPAIRED_2026-08-09.xlsx"

OUT_HUMAN   = SUMMARY / "workbook_human_prices_corrected.csv"
OUT_LLM     = SUMMARY / "workbook_llm_corrected.csv"
OUT_METRICS = SUMMARY / "workbook_metrics.csv"
OUT_README  = SUMMARY / "workbook_readme.md"
OUT_LOCKED  = WB_DIR  / "Master_Data_LOCKED_2026-08-13.xlsx"
OUT_CORRECTED = WB_DIR / "Master_Data_CORRECTED_2026-08-13.xlsx"

TODAY = "2026-08-13"

# ── Column indices (1-based) ──────────────────────────────────────────────────
# Human_Data_Entry
H_CO, H_YR, H_QTR, H_RATER = 1, 3, 4, 7
H_DOC_DATE = 12      # L  Document Date
H_M, H_N, H_O, H_P  = 13, 14, 15, 16  # Prior Closing Date, Prior Close, Next Opening Date, Next Day Open

# LLM_Data_Entry
L_CO, L_TK, L_YR, L_QTR = 1, 2, 3, 4
L_RATER = 5
L_SCORE, L_DEC = 7, 8       # G Sentiment, H Decision
L_TIME = 9                   # I Time
L_DOC  = 10                  # J Document Date
L_K, L_L, L_M, L_N = 11, 12, 13, 14   # Closing Date, Prior Close, Opening Date, Next Day Open
L_O_TOK = 15                 # O Token Cost
L_P_PCT = 16                 # P Actual % Change
L_Q_DIR = 17                 # Q Actual Direction
L_R_COR = 18                 # R Prediction Correct?
L_S_POS = 19                 # S Position
L_T_NET = 20                 # T Net P&L
L_U_NOT = 21                 # U Notes

COST_BPS      = 10    # Settings B4
SHORT_BORROW  = 0     # Settings B5
BAND          = 0.02  # ±2% grading band

# Exclusion set
WS_FLAGS  = SUMMARY / "worksheet_leak_flags.csv"
RET_MATRIX = SUMMARY / "returns_matrix.csv"
CAL_CSV    = SUMMARY / "global_outcome_calibration_phase2.csv"
COST_CSV   = SUMMARY / "api_cost_ledger.csv"
HUMAN_CSV  = ROOT / "data" / "human" / "human_decisions_export_2026-08-12.csv"
ABLATION   = SUMMARY / "section_ablation_results.csv"
ABLATION_COST = SUMMARY / "section_ablation_cost_per_correct.csv"
ABLATION_PAIRED = SUMMARY / "section_ablation_paired_diffs.csv"
HVL_STATS  = SUMMARY / "human_vs_llm_statistics.csv"
HVL_DECOMP = SUMMARY / "human_vs_llm_direction_decomposition.csv"
FRONTIER   = SUMMARY / "frontier_table.csv"
KAPPA_CSV  = SUMMARY / "kappa_near_independence.csv"
HOLDING_CURVE = SUMMARY / "ext2_holding_curve.csv"
WALKFORWARD_JSON = SUMMARY / "item_e_walkforward.json"
HUMAN_DEC_CSV = HUMAN_CSV


def load_csv_data(path, skip_hash=True):
    """Load CSV, skipping comment lines (starting with #)."""
    rows = []
    with open(path) as f:
        lines = [l for l in f if not (skip_hash and l.startswith('#'))]
    reader = csv.DictReader(lines)
    return list(reader)


def load_exclusion_set():
    ws = set()
    with open(WS_FLAGS) as f:
        for r in csv.DictReader(f):
            if r["has_worksheet"] == "True" and r["has_human_score"] == "True":
                ws.add(r["document_id"])
    spot = set(EXCLUDED_EVENTS)
    rm_rows = load_csv_data(RET_MATRIX)
    timing = {r["document_id"] for r in rm_rows if r["timing_excluded"] == "YES"}
    return ws | spot | timing


def load_returns_matrix():
    rows = load_csv_data(RET_MATRIX)
    return {r["document_id"]: r for r in rows}


def load_calibration():
    rows = load_csv_data(CAL_CSV)
    return {r["document_id"]: r for r in rows}


def load_cost_ledger():
    """Return {document_id: cost_usd} for micro layer only."""
    rows = load_csv_data(COST_CSV)
    costs = {}
    for r in rows:
        if r.get("layer") == "micro":
            did = r.get("document_id", "")
            try:
                costs[did] = float(r.get("estimated_cost_usd", 0))
            except ValueError:
                pass
    return costs


def load_ablation_by_event():
    """Return {document_id: {arm: row}} for section ablation."""
    rows = load_csv_data(ABLATION)
    result = defaultdict(dict)
    for r in rows:
        did = r.get("document_id", "")
        arm = r.get("arm", "")
        if did and arm:
            result[did][arm] = r
    return result


def next_bday(date_str):
    """Return date string of next business day after date_str."""
    dt = pd.Timestamp(date_str)
    nxt = dt + pd.tseries.offsets.BDay(1)
    return nxt.strftime("%Y-%m-%d")


def doc_id(ticker, year, quarter):
    """Construct document_id from ticker, year (int), quarter (str like Q1)."""
    qnum = str(quarter).replace("Q", "").replace("q", "")
    return f"{ticker}_FQ{qnum}_{year}"


def grading(ret):
    """Return (direction, graded). direction: UP/FLAT/DOWN on ±2% band."""
    if ret > BAND:
        return "UP", True
    elif ret < -BAND:
        return "DOWN", True
    else:
        return "FLAT", False


def prediction_correct(decision, direction):
    if decision == "BUY"  and direction == "UP":   return "YES"
    if decision == "SELL" and direction == "DOWN":  return "YES"
    if decision == "HOLD" and direction == "FLAT":  return "YES"
    return "NO"


def net_pnl(decision, ret):
    if decision == "BUY":
        pos = 1
    elif decision == "SELL":
        pos = -1
    else:
        return 0.0
    gross = pos * ret
    cost  = COST_BPS / 10000
    borrow = (SHORT_BORROW / 10000) if pos < 0 else 0
    return round(gross - cost - borrow, 6)


# ─────────────────────────────────────────────────────────────────────────────
# FILE 1: Human prices CSV
# ─────────────────────────────────────────────────────────────────────────────

def build_human_prices(wb, rm, excluded):
    ws = wb["Human_Data_Entry"]
    # Build company->ticker map from human_decisions_export
    co_ticker = {}  # (company_lower, year, quarter) -> ticker
    with open(HUMAN_CSV) as f:
        for row in csv.DictReader(f):
            co = row.get("company", "").strip()
            tk = row.get("ticker", "").strip()
            yr_str = row.get("year", "")
            qtr = row.get("quarter", "").strip()
            if co and tk and yr_str:
                try:
                    yr = int(yr_str)
                    co_ticker[(co.lower(), yr, qtr)] = tk
                except ValueError:
                    pass

    # Also add from LLM sheet
    ws_l = wb["LLM_Data_Entry"]
    for r in range(3, ws_l.max_row + 1):
        co = ws_l.cell(r, L_CO).value
        tk = ws_l.cell(r, L_TK).value
        yr = ws_l.cell(r, L_YR).value
        qtr = ws_l.cell(r, L_QTR).value
        if co and tk and yr and qtr:
            co_ticker[(str(co).lower(), int(yr), str(qtr))] = str(tk)

    out_rows = []
    n_llm = n_edgar = n_not = 0
    assert_failures = []

    for row_idx in range(3, ws.max_row + 1):
        co = ws.cell(row_idx, H_CO).value
        if not co:
            continue
        yr  = ws.cell(row_idx, H_YR).value
        qtr = ws.cell(row_idx, H_QTR).value
        rater = ws.cell(row_idx, H_RATER).value

        # Compute event_key
        event_key = f"{co}|{yr}|{qtr}"

        # Current (original) prices
        orig_M = ws.cell(row_idx, H_M).value
        orig_N = ws.cell(row_idx, H_N).value
        orig_O = ws.cell(row_idx, H_O).value
        orig_P = ws.cell(row_idx, H_P).value

        # Look up ticker
        tk = co_ticker.get((str(co).lower(), int(yr), str(qtr)))
        did = doc_id(tk, yr, qtr) if tk else None
        rm_row = rm.get(did) if did else None

        if rm_row:
            # Correctable from returns_matrix
            release_date = rm_row["report_date"]
            entry_date   = rm_row["entry_date"]
            entry_close  = float(rm_row["entry_close"])
            ret_ov       = float(rm_row["ret_overnight"])
            next_open    = round(entry_close * (1 + ret_ov), 2)

            # Timing: pre_market if entry_date < release_date, after_hours if equal
            if entry_date < release_date:
                timing = "pre_market"
                next_opening_date = release_date
            else:
                timing = "after_hours"
                next_opening_date = next_bday(entry_date)

            shifted = (str(orig_M.date()) if hasattr(orig_M, "date") else str(orig_M)) != entry_date

            out_rows.append({
                "row_index": row_idx,
                "event_key": event_key,
                "company": co,
                "ticker": tk,
                "year": yr,
                "quarter": qtr,
                "correctable": "TRUE",
                "prior_closing_date": entry_date,
                "prior_close": entry_close,
                "next_opening_date": next_opening_date,
                "next_day_open": next_open,
                "release_date": release_date,
                "release_timing": timing,
                "release_date_source": "returns_matrix.csv",
                "entry_shifted_vs_legacy": "TRUE" if shifted else "FALSE",
                "report_date_corrected": "YES",
                "original_prior_close": float(orig_N) if orig_N else "",
                "original_next_day_open": float(orig_P) if orig_P else "",
                "price_delta_flag": "YES" if shifted else "NO",
                "not_correctable_reason": "",
            })
            n_llm += 1
        else:
            # Human-only event
            reason = "human-only event; not in LLM returns_matrix — EDGAR lookup deferred, manual verification required"
            out_rows.append({
                "row_index": row_idx,
                "event_key": event_key,
                "company": co,
                "ticker": tk or "",
                "year": yr,
                "quarter": qtr,
                "correctable": "FALSE",
                "prior_closing_date": "",
                "prior_close": "",
                "next_opening_date": "",
                "next_day_open": "",
                "release_date": "",
                "release_timing": "",
                "release_date_source": "",
                "entry_shifted_vs_legacy": "",
                "report_date_corrected": "NO",
                "original_prior_close": float(orig_N) if orig_N else "",
                "original_next_day_open": float(orig_P) if orig_P else "",
                "price_delta_flag": "",
                "not_correctable_reason": reason,
            })
            n_not += 1

    print(f"  Human prices: {n_llm} from returns_matrix, {n_edgar} via EDGAR, {n_not} not correctable")

    fieldnames = [
        "row_index","event_key","company","ticker","year","quarter",
        "correctable","prior_closing_date","prior_close","next_opening_date","next_day_open",
        "release_date","release_timing","release_date_source","entry_shifted_vs_legacy",
        "report_date_corrected","original_prior_close","original_next_day_open",
        "price_delta_flag","not_correctable_reason",
    ]
    with open(OUT_HUMAN, "w", newline="") as f:
        w = csv.DictWriter(f, fieldnames=fieldnames)
        w.writeheader()
        w.writerows(out_rows)
    print(f"  Wrote {len(out_rows)} rows to {OUT_HUMAN.name}")
    return out_rows, assert_failures


# ─────────────────────────────────────────────────────────────────────────────
# FILE 2: LLM corrected CSV
# ─────────────────────────────────────────────────────────────────────────────

def build_llm_csv(wb, rm, cal, costs, excluded, ablation):
    ws = wb["LLM_Data_Entry"]

    # Section ablation four-arm set
    four_arm_dids = set()
    abl_rows = load_csv_data(ABLATION)
    for r in abl_rows:
        if r.get("document_id"):
            four_arm_dids.add(r["document_id"])
    # Two-arm set (full_bundle + press_release scored)
    two_arm_dids = set()
    for did, arms in ablation.items():
        if "full_bundle" in arms or "press_release" in arms:
            two_arm_dids.add(did)

    out_rows = []
    for row_idx in range(3, ws.max_row + 1):
        co = ws.cell(row_idx, L_CO).value
        if not co:
            continue
        tk  = ws.cell(row_idx, L_TK).value
        yr  = ws.cell(row_idx, L_YR).value
        qtr = ws.cell(row_idx, L_QTR).value

        event_key = f"{co}|{yr}|{qtr}"
        qnum = str(qtr).replace("Q","").replace("q","")
        did  = f"{tk}_FQ{qnum}_{yr}"

        rm_row  = rm.get(did, {})
        cal_row = cal.get(did, {})

        ret_ov  = float(rm_row["ret_overnight"]) if rm_row else 0.0
        entry_close = float(rm_row["entry_close"]) if rm_row else float(ws.cell(row_idx, L_L).value or 0)
        entry_date  = rm_row.get("entry_date", "")
        release_date = rm_row.get("report_date", "")
        next_open = round(entry_close * (1 + ret_ov), 2) if rm_row else 0.0

        # Timing
        if entry_date and release_date:
            if entry_date < release_date:
                timing = "pre_market"
                next_opening_date = release_date
            else:
                timing = "after_hours"
                next_opening_date = next_bday(entry_date)
        else:
            timing = ""
            next_opening_date = ""

        # LLM decision
        decision = cal_row.get("blend_predicted_signal_default", ws.cell(row_idx, L_DEC).value or "")
        score    = float(cal_row.get("micro_score", 0) or 0) * 0.55 + float(cal_row.get("macro_score", 0) or 0) * 0.45
        # Use blend score from calibration if possible
        blend_score = ws.cell(row_idx, L_SCORE).value  # original from workbook

        direction, graded = grading(ret_ov)
        correct = prediction_correct(decision, direction)
        pos = 1 if decision == "BUY" else (-1 if decision == "SELL" else 0)
        npnl = net_pnl(decision, ret_ov)
        tok_cost = costs.get(did, 0.0)

        is_excluded = did in excluded
        excl_reason = ""
        if is_excluded:
            ws_excl = set()
            with open(WS_FLAGS) as f:
                for r in csv.DictReader(f):
                    if r["has_worksheet"] == "True" and r["has_human_score"] == "True":
                        ws_excl.add(r["document_id"])
            if did in ws_excl:
                excl_reason = "worksheet contamination"
            elif did in EXCLUSION_REASONS:
                excl_reason = EXCLUSION_REASONS[did]
            elif rm_row.get("timing_excluded") == "YES":
                excl_reason = "timing unresolved (non-US issuer)"

        # Item C ablation
        abl = ablation.get(did, {})
        def arm_fields(arm_key):
            a = abl.get(arm_key, {})
            return {
                "decision": a.get("arm_signal", a.get("signal", "")),
                "correct": a.get("correct", ""),
                "tokens": a.get("total_tokens", ""),
            }
        fb = arm_fields("full_bundle")
        pr = arm_fields("press_release")
        pm = arm_fields("prepared_remarks")
        qa = arm_fields("qa_only")

        in_four_arm = did in four_arm_dids
        in_two_arm  = did in two_arm_dids

        out_rows.append({
            "row_index": row_idx,
            "event_key": event_key,
            "company": co,
            "ticker": tk,
            "year": yr,
            "quarter": qtr,
            "sentiment_score": blend_score,
            "decision": decision,
            "token_cost": round(tok_cost, 6),
            "run_id": rm_row.get("run_id", ""),
            "model_version": "deepseek-chat",
            "run_date": "",
            "document_date": release_date,
            "closing_date": entry_date,
            "prior_close": entry_close,
            "opening_date": next_opening_date,
            "next_day_open": next_open,
            "actual_pct_change": round(ret_ov * 100, 4),
            "actual_direction": direction,
            "prediction_correct": correct,
            "position": pos,
            "net_pnl": npnl,
            "release_date": release_date,
            "release_timing": timing,
            "release_date_source": "returns_matrix.csv" if rm_row else "",
            "excluded": "TRUE" if is_excluded else "FALSE",
            "exclusion_reason": excl_reason,
            "in_four_arm_set": "TRUE" if in_four_arm else "FALSE",
            "in_two_arm_set": "TRUE" if in_two_arm else "FALSE",
            "arm_full_decision": fb["decision"],
            "arm_full_correct": fb["correct"],
            "arm_full_tokens": fb["tokens"],
            "arm_pr_decision": pr["decision"],
            "arm_pr_correct": pr["correct"],
            "arm_pr_tokens": pr["tokens"],
            "arm_prepared_decision": pm["decision"],
            "arm_prepared_correct": pm["correct"],
            "arm_prepared_tokens": pm["tokens"],
            "arm_qa_decision": qa["decision"],
            "arm_qa_correct": qa["correct"],
            "arm_qa_tokens": qa["tokens"],
            "notes": ws.cell(row_idx, L_U_NOT).value or "",
        })

    fieldnames = [
        "row_index","event_key","company","ticker","year","quarter",
        "sentiment_score","decision","token_cost","run_id","model_version","run_date",
        "document_date","closing_date","prior_close","opening_date","next_day_open",
        "actual_pct_change","actual_direction","prediction_correct","position","net_pnl",
        "release_date","release_timing","release_date_source",
        "excluded","exclusion_reason",
        "in_four_arm_set","in_two_arm_set",
        "arm_full_decision","arm_full_correct","arm_full_tokens",
        "arm_pr_decision","arm_pr_correct","arm_pr_tokens",
        "arm_prepared_decision","arm_prepared_correct","arm_prepared_tokens",
        "arm_qa_decision","arm_qa_correct","arm_qa_tokens",
        "notes",
    ]
    with open(OUT_LLM, "w", newline="") as f:
        w = csv.DictWriter(f, fieldnames=fieldnames)
        w.writeheader()
        w.writerows(out_rows)
    print(f"  Wrote {len(out_rows)} rows to {OUT_LLM.name}")
    return out_rows


# ─────────────────────────────────────────────────────────────────────────────
# FILE 3: Metrics CSV
# ─────────────────────────────────────────────────────────────────────────────

def build_metrics_csv():
    rows = []

    def add(name, value, unit="", n="", denom="", event_set="", convention="",
            ci_low="", ci_high="", ci_level="", p_value="", mde="",
            test="", threshold_dep="FALSE", source="", notes=""):
        rows.append({
            "metric_name": name, "value": value, "unit": unit, "n": n,
            "denominator_definition": denom, "event_set": event_set,
            "convention": convention, "ci_low": ci_low, "ci_high": ci_high,
            "ci_level": ci_level, "p_value": p_value, "mde": mde,
            "test_used": test, "threshold_dependent": threshold_dep,
            "source_file": source, "notes": notes,
        })

    # ── Effective sample funnel ──────────────────────────────────────────────
    add("total_events_scored", 268, "events", n=268,
        event_set="N=268 phase2 scored events", source="effective_sample_funnel.md")
    add("excluded_worksheet", 25, "events", n=268,
        event_set="N=268", source="worksheet_leak_flags.csv",
        notes="LLM input included human rater worksheet")
    add("excluded_bad_document", len(EXCLUDED_EVENTS), "events",
        source="eval/excluded_events.py",
        notes="; ".join(sorted(EXCLUDED_EVENTS)))
    add("excluded_timing", 9, "events",
        source="effective_sample_funnel.md", notes="9 non-US issuers, timing unresolved")
    # 233 until DIS_FQ1_2025 was excluded on 2026-08-24 for look-ahead.
    add("clean_universe", 268 - 25 - len(EXCLUDED_EVENTS) - 9, "events",
        event_set="N=232 clean universe", source="eval/excluded_events.py")
    add("traded_events", 146, "events",
        denom="Model called BUY or SELL",
        event_set="N=233 clean universe", threshold_dep="TRUE", source="ext2_holding_curve.csv")
    add("graded_events", 95, "events",
        denom="|ret_overnight| > 2% AND traded",
        event_set="N=233 clean universe",
        convention="pre-registered ±2% overnight band",
        threshold_dep="TRUE", source="ext2_holding_curve.csv")

    # ── Selectivity accuracy ─────────────────────────────────────────────────
    add("selectivity_accuracy", "65.3%", "percent",
        n=95, denom="model traded AND |ret_overnight|>2%",
        event_set="N=233 clean universe",
        convention="HOLD-excluded, pre-registered ±2% band",
        ci_low="", ci_high="", p_value="0.024", mde="±10.8pp",
        test="binomial vs always-DOWN floor 54.7%",
        threshold_dep="TRUE", source="ext2_holding_curve.csv",
        notes="62/95; margin +10.5pp vs floor 54.7% (52/95 always-DOWN)")
    add("selectivity_accuracy_numerator", "62/95", "",
        n=95, event_set="N=233 clean universe",
        threshold_dep="TRUE", source="ext2_holding_curve.csv")
    add("selectivity_floor_always_down", "54.7%", "percent",
        n=95, denom="always-DOWN (predict SELL every event) on graded set",
        event_set="N=233 clean universe",
        convention="52/95 graded events have negative return",
        source="ext2_holding_curve.csv")
    add("selectivity_margin_vs_floor", "+10.5pp", "pp",
        n=95, p_value="0.024", mde="±10.8pp",
        test="binomial one-sided greater vs always-DOWN 52/95",
        threshold_dep="TRUE", source="ext2_holding_curve.csv")

    # ── Coverage accuracy ────────────────────────────────────────────────────
    add("coverage_accuracy", "42.2%", "percent",
        n=147, denom="all events with |ret_overnight|>2% (model traded or not)",
        event_set="N=233 clean universe",
        convention="HOLD=wrong, pre-registered ±2% band",
        source="ext2_holding_curve.csv",
        notes="62/147; floor=54.4% (80/147 always-DOWN); margin -12.2pp. "
              "Model missed 52 events that moved.")
    add("coverage_floor_always_down", "54.4%", "percent",
        n=147, denom="all events with |ret_overnight|>2%",
        event_set="N=233 clean universe",
        notes="80/147 events have negative return")

    # ── P&L metrics ──────────────────────────────────────────────────────────
    hc = load_csv_data(HOLDING_CURVE)
    hc_map = {r["horizon"]: r for r in hc}
    ov = hc_map.get("overnight", {})

    add("mean_net_per_trade_overnight", f"{float(ov.get('mean_net_per_trade',0))*100:.4f}%", "percent",
        n=146, denom="traded events",
        event_set="N=233 clean universe",
        convention="overnight, 10bps round-trip cost, 0bps short borrow",
        ci_low=f"{float(ov.get('bootstrap_ci_low',0))*100:.2f}%",
        ci_high=f"{float(ov.get('bootstrap_ci_high',0))*100:.2f}%",
        ci_level="90%",
        threshold_dep="TRUE", source="ext2_holding_curve.csv",
        notes="+1.862% per trade")

    # Summed total return — from backtest_equity.csv
    try:
        beq = load_csv_data(SUMMARY / "backtest_equity.csv")
        beq_map = {r.get("metric",r.get("statistic","")): r.get("value","") for r in beq}
    except Exception:
        beq_map = {}

    add("summed_total_return", "+271.81%", "percent (order-dependent)",
        n=146,
        event_set="N=233 clean universe — N=268 OLD ANCHOR",
        notes="Order-dependent compounded equity curve, not an achievable balance. "
              "Mean net per trade is the primary P&L metric. "
              "Value from item_e_handoff.md (+271.81% on N=233).",
        source="item_e_handoff.md", threshold_dep="TRUE")

    add("t_statistic_per_trade", "3.43", "",
        n=146, denom="traded events",
        event_set="N=233 clean universe",
        convention="mean_net/pstdev(nets)*sqrt(N_traded) — NOT a time-series Sharpe",
        source="item_e_handoff.md", threshold_dep="TRUE",
        notes="Grows mechanically with sample size. Do not label as Sharpe.")

    add("info_ratio_per_trade", "0.284", "",
        n=146, source="item_e_handoff.md", threshold_dep="TRUE",
        notes="mean_net / pstdev(nets), unscaled by sqrt(N)")

    # ── Rank correlations by horizon ─────────────────────────────────────────
    for row in hc:
        hz = row["horizon"]
        rho = float(row["rank_correlation"])
        p   = float(row["rho_pvalue"])
        band_status = row.get("band_status","")
        add(f"spearman_rho_{hz}", f"{rho:.4f}", "",
            n=233 if hz=="overnight" else 233,
            event_set="N=233 clean universe",
            convention=f"{hz} return, {row.get('hold_band','')} band ({band_status})",
            p_value=f"{p:.4f}",
            test="Spearman rank correlation, blended score vs ret_overnight",
            source="ext2_holding_curve.csv")
        add(f"mean_net_per_trade_{hz}", f"{float(row['mean_net_per_trade'])*100:.4f}%", "percent",
            n=int(row["traded_n"]),
            event_set="N=233 clean universe",
            ci_low=f"{float(row['bootstrap_ci_low'])*100:.4f}%",
            ci_high=f"{float(row['bootstrap_ci_high'])*100:.4f}%",
            ci_level="90%",
            source="ext2_holding_curve.csv")
        add(f"accuracy_{hz}", f"{float(row['accuracy'])*100:.1f}%", "percent",
            n=int(row["graded_n"]), denom=f"|ret_{hz}|>band",
            source="ext2_holding_curve.csv", threshold_dep="TRUE")

    # ── Majority-direction floor (full sample) ───────────────────────────────
    add("majority_direction_floor_full_sample", "54.7%", "percent",
        n=95, denom="always-DOWN on graded events (|ret_overnight|>2%, traded)",
        event_set="N=233 clean universe, graded events only",
        convention="52/95 graded events have negative return",
        source="ext2_holding_curve.csv")

    # ── Kappa ────────────────────────────────────────────────────────────────
    kappa_rows = load_csv_data(KAPPA_CSV)
    kappa_map = {r["metric"]: r["value"] for r in kappa_rows if "metric" in r}
    add("cohens_kappa_human_vs_llm", kappa_map.get("cohens_kappa","0.109"), "",
        n=int(kappa_map.get("n_paired_events", 171)),
        denom="paired events with both human and LLM BUY/HOLD/SELL call",
        event_set="N=171 paired events (section=All, first_rater=YES, in_llm=YES, exclusions applied)",
        convention="3x3 BUY/HOLD/SELL vs BUY/HOLD/SELL, no return grading",
        ci_low=kappa_map.get("bootstrap_90ci_low","0.030"),
        ci_high=kappa_map.get("bootstrap_90ci_high","0.190"),
        ci_level="90%",
        test="bootstrap 10000 resamples, seed 20260709",
        source="kappa_near_independence.csv")

    # ── Human vs LLM marginals ───────────────────────────────────────────────
    hvl = {r["statistic"]: r["value"] for r in load_csv_data(HVL_STATS)}
    for arm in ["human","llm"]:
        for call in ["buy","hold","sell"]:
            k = f"{arm}_call_{call}_count"
            kf = f"{arm}_call_{call}_frac"
            add(f"call_count_{arm}_{call}", hvl.get(k,""), "count",
                n=171, event_set="N=171 paired events",
                source="human_vs_llm_statistics.csv")
            add(f"call_frac_{arm}_{call}", hvl.get(kf,""), "fraction",
                n=171, source="human_vs_llm_statistics.csv")

    # ── Direction-only comparison ────────────────────────────────────────────
    add("direction_only_n", hvl.get("direction_only_n","76"), "events",
        event_set="N=76 events where both arms called BUY or SELL",
        source="human_vs_llm_statistics.csv")
    add("direction_human_sign_accuracy", hvl.get("direction_human_sign_accuracy",""),
        "fraction", n=76,
        event_set="N=76 direction-only",
        convention="sign accuracy, no band filter",
        source="human_vs_llm_statistics.csv", notes="44/76")
    add("direction_llm_sign_accuracy", hvl.get("direction_llm_sign_accuracy",""),
        "fraction", n=76,
        source="human_vs_llm_statistics.csv", notes="46/76")
    add("direction_llm_minus_human_pp", hvl.get("direction_llm_minus_human_pp",""),
        "pp", n=76,
        ci_low=hvl.get("direction_90ci_low_pp",""),
        ci_high=hvl.get("direction_90ci_high_pp",""),
        ci_level="90%",
        p_value=hvl.get("direction_p_value",""),
        test="bootstrap paired 10000 resamples seed 20260709",
        source="human_vs_llm_statistics.csv")

    # Both floor definitions for direction-only
    add("direction_floor_always_down_strict", "55.3%", "percent",
        n=76, denom="42/76 events with ret_overnight < 0 (strict negative)",
        event_set="N=76 direction-only",
        convention="ret < 0 strictly; excludes PUM.DE_FQ4_2025 (ret=0)",
        source="human_vs_llm_statistics.csv")
    add("direction_floor_always_down_incl_zero", "56.6%", "percent",
        n=76, denom="43/76 events with ret_overnight <= 0 (includes zero-return PUM.DE_FQ4_2025)",
        event_set="N=76 direction-only",
        convention="ret <= 0; used for per-direction base-rate in finding #7",
        source="human_vs_llm_statistics.csv",
        notes="PUM.DE_FQ4_2025 has ret_overnight=0.0 exactly")
    add("direction_human_vs_always_down_p", hvl.get("direction_human_vs_always_down_p",""),
        "", n=76,
        p_value=hvl.get("direction_human_vs_always_down_p",""),
        test="binomial one-sided greater vs 42/76=55.3%",
        source="human_vs_llm_statistics.csv")
    add("direction_llm_vs_always_down_p", hvl.get("direction_llm_vs_always_down_p",""),
        "", n=76,
        p_value=hvl.get("direction_llm_vs_always_down_p",""),
        test="binomial one-sided greater vs 42/76=55.3%",
        source="human_vs_llm_statistics.csv")

    # ── Per-direction accuracy (finding #7) ──────────────────────────────────
    decomp = load_csv_data(HVL_DECOMP)
    for row in decomp:
        arm  = row["arm"]
        call = row["call"]
        if call == "Overall":
            continue
        add(f"direction_{arm}_{call.lower()}_accuracy",
            f"{float(row['accuracy'])*100:.1f}%", "percent",
            n=int(row["n_total"]),
            denom=f"{arm} {call} calls among N=76 direction-only events",
            event_set="N=76 direction-only",
            convention="sign accuracy, no band",
            p_value=row["vs_50pct_p"],
            test=f"binomial one-sided greater vs 50%; vs base-rate p={row['vs_base_rate_p']} (base_rate={row['base_rate']})",
            source="human_vs_llm_direction_decomposition.csv",
            notes=f"{row['n_correct']}/{row['n_total']}, vs_50pct_margin={row['vs_50pct_margin_pp']}pp")

    # ── Item C ablation — DATA ROWS ONLY from section_ablation_cost_per_correct.csv ──
    abl_cost = load_csv_data(ABLATION_COST)
    for row in abl_cost:
        event_set_str = row["event_set"]
        arm = row["arm"]
        add(f"item_c_{event_set_str}_{arm}_accuracy",
            f"{float(row['accuracy_hold_excluded'])*100:.1f}%", "percent",
            n=int(row["n_graded"]),
            denom=f"graded events (|ret|>2%) in {event_set_str}",
            event_set=event_set_str,
            convention="HOLD-excluded, ±2% overnight band",
            threshold_dep="TRUE",
            source="section_ablation_cost_per_correct.csv",
            notes=f"{row['n_correct']}/{row['n_graded']}; calls={row['n_calls']}; mean_tokens={row['mean_tokens_per_call']}")
        add(f"item_c_{event_set_str}_{arm}_cost_per_correct",
            f"${float(row['cost_per_correct_usd']):.4f}", "USD per correct call",
            n=int(row["n_correct"]),
            event_set=event_set_str,
            source="section_ablation_cost_per_correct.csv",
            threshold_dep="TRUE",
            notes=f"total_cost=${float(row['total_cost_usd']):.4f}; total_tokens={row['total_tokens']}")
        add(f"item_c_{event_set_str}_{arm}_mean_tokens",
            row["mean_tokens_per_call"], "tokens per call",
            event_set=event_set_str,
            source="section_ablation_cost_per_correct.csv")

    # Item C paired diffs
    abl_paired = load_csv_data(ABLATION_PAIRED)
    for row in abl_paired:
        name = f"item_c_paired_{row.get('comparison','').replace(' vs ','_vs_').replace(' ','_')}_{row.get('method','').replace(' ','_').replace(',','')}"
        add(name,
            row.get("point_diff",""), "fraction diff",
            n=int(row.get("n_paired",0)),
            event_set=row.get("set","four_arm"),
            convention=row.get("method",""),
            ci_low=row.get("ci_low",""),
            ci_high=row.get("ci_high",""),
            p_value=row.get("p_value",""),
            source="section_ablation_paired_diffs.csv",
            threshold_dep="TRUE",
            notes=row.get("note",""))

    # ── FinBERT frontier ──────────────────────────────────────────────────────
    frontier_rows = load_csv_data(FRONTIER)
    for row in frontier_rows:
        method = row.get("method","")
        add(f"frontier_{method.replace(' ','_').replace(',','').replace('/','_')}_accuracy_flat_excluded",
            f"{float(row.get('accuracy_flat_excluded',0))*100:.1f}%", "percent",
            n=int(row.get("graded_n_flat_excluded",0)),
            denom="graded events, FLAT-excluded",
            event_set="eval split (186 events, 119 graded)",
            ci_low=row.get("bootstrap_ci_flat_excluded","").strip("[]").split(",")[0] if "[" in row.get("bootstrap_ci_flat_excluded","") else "",
            ci_high=row.get("bootstrap_ci_flat_excluded","").strip("[]").split(",")[-1] if "[" in row.get("bootstrap_ci_flat_excluded","") else "",
            ci_level="90%",
            source="frontier_table.csv",
            notes=row.get("note",""))
        add(f"frontier_{method.replace(' ','_').replace(',','').replace('/','_')}_accuracy_flat_wrong",
            f"{float(row.get('accuracy_flat_as_wrong',0))*100:.1f}%", "percent",
            n=int(row.get("graded_n_flat_as_wrong",0)),
            denom="all eval events, HOLD=wrong",
            event_set="eval split (186 events)",
            source="frontier_table.csv")

    # ── Walk-forward results ──────────────────────────────────────────────────
    with open(WALKFORWARD_JSON) as f:
        wf = json.load(f)

    rw = wf.get("rolling_walkforward", {})
    add("walkforward_n_windows", str(rw.get("n_windows","")), "",
        source="item_e_walkforward.json")
    add("walkforward_degeneracy_finding",
        rw.get("degeneracy_finding","degenerates at N=233"), "",
        source="item_e_walkforward.json",
        notes="Both mean-net and accuracy objectives tried; both degenerate")

    for obj_key in ["mean_net_objective","accuracy_objective"]:
        obj = rw.get(obj_key, {})
        add(f"walkforward_{obj_key}_degenerate_windows",
            str(obj.get("n_degenerate_windows", "")), "",
            event_set=f"walk-forward {obj_key}",
            source="item_e_walkforward.json",
            notes=str(obj.get("note","")))

    # ── Dev/eval split ────────────────────────────────────────────────────────
    isd = wf.get("in_sample_deployed", {})
    add("in_sample_deployed_accuracy", f"{float(isd.get('accuracy',0))*100:.1f}%", "percent",
        n=int(isd.get("n_graded",0)),
        denom="graded events, deployed thresholds, full N=233",
        event_set="N=233 clean universe",
        threshold_dep="TRUE", source="item_e_walkforward.json",
        notes=f"{isd.get('n_correct','')}/{isd.get('n_graded','')}; trades={isd.get('n_trades','')}")

    # Eval split floor
    add("eval_split_always_down_floor", "54.6%", "percent",
        n=119, denom="65/119 graded eval events have negative return",
        event_set="eval split (latest 80% by release_date, 186 events, 119 graded)",
        source="frontier_table.csv")
    add("eval_split_accuracy", "68.0%", "percent",
        n=75, denom="traded+graded events in eval split",
        event_set="eval split, N=186",
        p_value="0.013", mde="±14.4pp",
        test="binomial vs always-DOWN floor 54.6%",
        threshold_dep="TRUE", source="surviving_findings.md",
        notes="51/75; margin +13.4pp; sorted by release_date, latest 80%")
    add("dev_split_accuracy", "55.0%", "percent",
        n=20, denom="traded+graded events in dev split",
        event_set="dev split, N=47 (earliest 20% by release_date)",
        mde="±27.8pp",
        threshold_dep="TRUE", source="surviving_findings.md",
        notes="11/20; margin +1.4pp vs floor 53.6% — not significant, underpowered")

    # ── Cross-issuer generalisation ───────────────────────────────────────────
    gu = wf.get("genuinely_unseen", {})
    ps = gu.get("post_sweep", {})
    add("cross_issuer_accuracy", f"{float(ps.get('accuracy',0))*100:.1f}%", "percent",
        n=int(ps.get("n_graded",0)),
        denom="graded events in 29 post-sweep issuers",
        event_set="N=101 post-sweep events, 29 issuers (2 further timing-excluded)",
        p_value=str(ps.get("binomial_p","")),
        mde=f"±{ps.get('mde_pp','')}pp",
        test="binomial vs always-DOWN floor",
        threshold_dep="TRUE", source="item_e_walkforward.json",
        notes=f"{ps.get('n_correct','')}/{ps.get('n_graded','')}; floor={float(ps.get('majority_direction_floor',0))*100:.1f}%; "
              "RECONSTRUCTED subset — sweep membership inferred from issuer order; "
              "cross-issuer test only, NOT temporal generalisation; "
              "dates overlap with in-sweep events (both span 2023-2026)")

    fields = [
        "metric_name","value","unit","n","denominator_definition","event_set","convention",
        "ci_low","ci_high","ci_level","p_value","mde","test_used","threshold_dependent",
        "source_file","notes",
    ]
    with open(OUT_METRICS, "w", newline="") as f:
        w = csv.DictWriter(f, fieldnames=fields)
        w.writeheader()
        w.writerows(rows)
    print(f"  Wrote {len(rows)} rows to {OUT_METRICS.name}")


# ─────────────────────────────────────────────────────────────────────────────
# FILE 4: README
# ─────────────────────────────────────────────────────────────────────────────

def build_readme():
    text = """# Workbook correction notes

## Original price basis (superseded)

The source workbook `Master_Data_NEW_REPAIRED_2026-08-09.xlsx` uses the close of
`report_date` as the entry price uniformly for all events. This convention is wrong
for approximately 82 pre_market events: for those events, the 8-K Item 2.02 was filed
before market open, so the relevant entry price is the open on `release_date` (the filing
date), not the close on that same day. Using the close-on-release_date convention for
pre_market events inflates the measured overnight move, because the close already reflects
the announced results.

## Section C re-pricing block (invalid)

The Section C re-pricing section of the source workbook selected measurement windows by
observed price movement (choosing whichever window showed the largest move). This is
selection bias: windows are chosen because they moved, not because they are the correct
entry/exit point. Any accuracy or P&L figures derived from Section C are not valid and
must not be used.

## Corrected basis

The corrected entry anchor is `release_date` from the EDGAR 8-K Item 2.02 filing date,
with the following timing rule:

- **pre_market** events (8-K filed before market open): entry close = close on the
  preceding trading day; entry open = open on `release_date`.
- **after_hours** events (8-K filed after market close): entry close = close on
  `release_date`; entry open = open on the next trading day.

The timing classification is inferred from the relationship between `entry_date` and
`report_date` in `returns_matrix.csv`: if `entry_date < report_date` then pre_market;
if `entry_date == report_date` then after_hours.

**Date correction established**: 2026-08-12. Source: `item_e_handoff.md`,
`returns_matrix.csv`, `retracted_findings_2026-08-12.md`.

## What changed

- **`Master_Data_LOCKED_2026-08-13.xlsx`**: Human_Data_Entry prices unchanged.
  LLM_Data_Entry updated to corrected figures. New sheets appended. A banner in
  Human_Data_Entry A1 notes that the human arm prices remain on the superseded basis.

- **`Master_Data_CORRECTED_2026-08-13.xlsx`**: Human_Data_Entry prices corrected
  for all events where `correctable=TRUE` in `workbook_human_prices_corrected.csv`.
  Events marked `correctable=FALSE` (human-only events not in the LLM returns_matrix)
  retain original prices; manual verification against EDGAR is required. LLM_Data_Entry
  updated. New sheets appended.

After opening either file in Excel, force a full recalculation with **Ctrl+Alt+F9**
before circulating. LibreOffice cannot evaluate XLOOKUP and will show #NAME? errors —
this does not indicate file corruption.
"""
    with open(OUT_README, "w") as f:
        f.write(text)
    print(f"  Wrote {OUT_README.name}")


# ─────────────────────────────────────────────────────────────────────────────
# Workbook helpers
# ─────────────────────────────────────────────────────────────────────────────

BOLD = Font(bold=True)

def make_sheet(wb, title, headers, rows_data, freeze="A2"):
    """Append a new sheet with bold header, freeze panes, auto-width."""
    ws = wb.create_sheet(title=title)
    ws.append(headers)
    for cell in ws[1]:
        cell.font = BOLD
    for row in rows_data:
        ws.append(row)
    ws.freeze_panes = freeze

    # Auto-width (estimate)
    col_widths = [len(str(h)) for h in headers]
    for row in ws.iter_rows(min_row=2):
        for i, cell in enumerate(row):
            if i < len(col_widths):
                col_widths[i] = max(col_widths[i], min(len(str(cell.value or "")), 60))
    for i, w in enumerate(col_widths):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i+1)].width = w + 2
    return ws


def build_corrections_log_sheet(wb, human_rows, llm_rows_data):
    """Build Corrections_Log sheet."""
    corrections = [
        ["What was claimed", "Corrected value", "Cause", "Date", "Affected events", "Source file"],
        ["report_date close as entry for all events",
         "release_date with timing rule (pre_market: prev close + open on release; after_hours: close on release + open next day)",
         "report_date ≠ release_date for ~50 US events; pre_market close was post-announcement",
         "2026-08-12", "~82 pre_market events", "retracted_findings_2026-08-12.md"],
        ["~39% always-BUY baseline",
         "54.7% always-DOWN (52/95 graded events negative)",
         "Wrong denominator: used all events including HOLDs instead of graded events only",
         "2026-08-13", "All accuracy comparisons", "baseline_correction_2026-08-13.md"],
        ["'Per-direction accuracy equals base rate in subset' claimed as finding",
         "Retracted: tautological (accuracy and subset base rate are the same number by definition)",
         "Definitional error — not a finding",
         "2026-08-13", "0 (not a data correction)", "direction_accuracy_decomposition.md"],
        ["72.7% band capture, 61% sign-correct p=0.064, PM/AH divergence, +17.3pp agreement filter",
         "42.4% band capture, 56% sign-correct p=0.480, no divergence, -0.3pp agreement filter",
         "All four derived from report_date anchor artefact in pre_market events",
         "2026-08-12", "82 pre_market events", "retracted_findings_2026-08-12.md"],
        ["Item C: PR 63.3% at $0.017/correct, full bundle 67.7% at $0.027/correct",
         "Four-arm N=119: PR 64.9% (24/37) at $0.0225; full 68.1% (32/47) at $0.0358. All-events N=200: PR 62.5% (35/56) at $0.029; full 65.7% (46/70) at $0.046",
         "Prose figures existed only in CSV comment headers; not backed by data rows",
         "2026-08-13", "All Item C accuracy/cost citations", "section_ablation_cost_per_correct.csv"],
        ["'31 issuers' for cross-issuer generalisation subset",
         "29 issuers with clean events in N=233 (2 further post-sweep issuers — Allianz, Lenovo — are fully timing-excluded)",
         "Count included timing-excluded issuers",
         "2026-08-13", "1 count citation", "item_e_walkforward.json"],
    ]

    ba_headers = ["Metric", "Basis", "Accuracy", "Mean net per trade", "Total return (order-dep.)", "N", "Source"]
    ba_rows = [
        # Model arm
        ["Model arm", "Original report_date (superseded)",
         "62.6% (N=268, 5-day calibration window, not overnight)",
         "+1.719% overnight", "+1243.64% (N=268, 146 trades)",
         "268 total / 171 traded / 92 graded", "CLAUDE.md historical, backtest_equity.csv"],
        ["Model arm", "Corrected release_date (N=233 clean)",
         "65.3% overnight (62/95 graded)",
         "+1.862% overnight", "+271.81% (N=233, 146 trades)",
         "233 clean / 146 traded / 95 graded", "item_e_handoff.md, ext2_holding_curve.csv"],
        # Human arm
        ["Human arm", "Original report_date (superseded, in LOCKED workbook)",
         "Recalculated by Excel formula from original M/N/O/P prices",
         "Recalculated by Excel formula",
         "Recalculated by Excel formula",
         "420 rows in Human_Data_Entry", "Master_Data_LOCKED_2026-08-13.xlsx"],
        ["Human arm", "Corrected release_date (in CORRECTED workbook)",
         "Recalculated by Excel formula from corrected M/N/O/P prices",
         "Recalculated by Excel formula",
         "Recalculated by Excel formula",
         f"{sum(1 for r in human_rows if r.get('correctable')=='TRUE')} rows corrected; {sum(1 for r in human_rows if r.get('correctable')=='FALSE')} left on original basis",
         "Master_Data_CORRECTED_2026-08-13.xlsx; workbook_human_prices_corrected.csv"],
    ]

    ws = wb.create_sheet(title="Corrections_Log")
    ws.append(["CORRECTIONS TABLE"])
    ws[1][0].font = BOLD
    ws.append([])
    for row in corrections:
        ws.append(row)
    for cell in ws[2]:
        cell.font = BOLD
    ws.append([])
    ws.append(["BEFORE / AFTER PERFORMANCE"])
    ws[ws.max_row][0].font = BOLD
    ws.append(ba_headers)
    for cell in ws[ws.max_row]:
        cell.font = BOLD
    for row in ba_rows:
        ws.append(row)
    ws.freeze_panes = "A3"
    for i in range(1, 8):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = 40
    return ws


def build_findings_sheet(wb):
    """Build Findings sheet from committed markdown files."""
    ws = wb.create_sheet(title="Findings")
    ws.append(["Finding", "Text"])
    for cell in ws[1]:
        cell.font = BOLD
    ws.freeze_panes = "A2"

    # Surviving findings
    surviving_path = SUMMARY / "surviving_findings.md"
    try:
        content = surviving_path.read_text()
        # Extract numbered headings
        sections = re.split(r'\n## ', content)
        for sec in sections[1:]:  # skip intro
            lines = sec.strip().split("\n")
            title = lines[0].strip()
            body = "\n".join(lines[1:]).strip()[:2000]
            ws.append([f"SURVIVING: {title}", body])
    except Exception as e:
        ws.append(["SURVIVING: (read error)", str(e)])

    ws.append([])
    ws.append(["RETRACTED FINDINGS", ""])

    # Retracted findings
    retracted_path = SUMMARY / "retracted_findings_2026-08-12.md"
    try:
        content = retracted_path.read_text()
        sections = re.split(r'\n## ', content)
        for sec in sections[1:]:
            lines = sec.strip().split("\n")
            title = lines[0].strip()
            body = "\n".join(lines[1:]).strip()[:1000]
            ws.append([f"RETRACTED: {title}", body])
    except Exception as e:
        ws.append(["RETRACTED: (read error)", str(e)])

    ws.column_dimensions["A"].width = 40
    ws.column_dimensions["B"].width = 80
    return ws


def build_confusion_matrices_sheet(wb, rm, excluded):
    """Build Confusion_Matrices sheet with three tabulations."""
    ws = wb.create_sheet(title="Confusion_Matrices")
    bold = BOLD

    def add_header(text):
        ws.append([])
        ws.append([text])
        ws[ws.max_row][0].font = bold
        ws.append([])

    # ── (a) Model call vs realised outcome, N=233 ────────────────────────────
    add_header("(a) Model call vs realised outcome — N=233 clean events, pre-registered ±2% overnight band")
    ws.append(["", "Outcome UP (ret>2%)", "Outcome FLAT (|ret|≤2%)", "Outcome DOWN (ret<-2%)", "Row total",
               "Precision (correct/called_dir)", "Recall (correct/realised_dir)"])
    ws[ws.max_row][0].font = bold

    cal = load_calibration()
    clean_ids = set(rm.keys()) - excluded
    calls = {"BUY": defaultdict(int), "HOLD": defaultdict(int), "SELL": defaultdict(int)}
    outcomes = {"UP": 0, "FLAT": 0, "DOWN": 0}
    for did in clean_ids:
        rm_r = rm.get(did, {})
        cal_r = cal.get(did, {})
        if not rm_r or not cal_r:
            continue
        ret = float(rm_r.get("ret_overnight", 0))
        direction, _ = grading(ret)
        decision = cal_r.get("blend_predicted_signal_default", "HOLD")
        calls[decision][direction] += 1
        outcomes[direction] += 1

    for call_label in ["BUY", "HOLD", "SELL"]:
        c = calls[call_label]
        row_total = sum(c.values())
        # Precision: correct/row_total
        if call_label == "BUY":
            correct_for_precision = c.get("UP", 0)
        elif call_label == "SELL":
            correct_for_precision = c.get("DOWN", 0)
        else:
            correct_for_precision = c.get("FLAT", 0)
        prec = f"{correct_for_precision}/{row_total}" if row_total else ""
        ws.append([
            f"Call {call_label}",
            c.get("UP", 0), c.get("FLAT", 0), c.get("DOWN", 0),
            row_total, prec, ""
        ])

    # Column totals and recall
    col_totals = [outcomes.get("UP", 0), outcomes.get("FLAT", 0), outcomes.get("DOWN", 0)]
    correct_up   = calls["BUY"].get("UP", 0)
    correct_flat = calls["HOLD"].get("FLAT", 0)
    correct_down = calls["SELL"].get("DOWN", 0)
    ws.append(["Col total"] + col_totals + [sum(col_totals), "", ""])
    ws.append(["Recall (correct/total_in_direction)",
               f"{correct_up}/{col_totals[0]}" if col_totals[0] else "",
               f"{correct_flat}/{col_totals[1]}" if col_totals[1] else "",
               f"{correct_down}/{col_totals[2]}" if col_totals[2] else "", "", "", ""])
    ws.append(["Note: BUY-called-DOWN and SELL-called-UP are the costly directional errors. "
               "HOLD row against UP and DOWN is the coverage gap (52 events that moved but were passed on).", "", "", "", "", "", ""])

    # ── (b) Human arm confusion matrix (N=171 paired) ────────────────────────
    add_header("(b) Human arm vs realised outcome — N=171 paired events (section=All, first_rater=YES, in_llm=YES, exclusions applied)")
    ws.append(["", "Outcome UP (ret>2%)", "Outcome FLAT (|ret|≤2%)", "Outcome DOWN (ret<-2%)", "Row total"])
    ws[ws.max_row][0].font = bold

    human_calls = {"BUY": defaultdict(int), "HOLD": defaultdict(int), "SELL": defaultdict(int)}
    with open(HUMAN_CSV) as f:
        for row in csv.DictReader(f):
            if row.get("section") != "All": continue
            if row.get("first_rater_for_event") != "YES": continue
            if row.get("in_llm_universe") != "YES": continue
            co = row.get("company",""); yr_str = row.get("year",""); qtr = row.get("quarter","")
            tk = row.get("ticker","")
            try:
                yr = int(yr_str)
            except ValueError:
                continue
            qnum = qtr.replace("Q","")
            did = f"{tk}_FQ{qnum}_{yr}"
            if did not in clean_ids:
                continue
            rm_r = rm.get(did, {})
            if not rm_r:
                continue
            ret = float(rm_r.get("ret_overnight", 0))
            direction, _ = grading(ret)
            dec = row.get("human_decision","").strip().upper()
            if dec in human_calls:
                human_calls[dec][direction] += 1

    for call_label in ["BUY", "HOLD", "SELL"]:
        c = human_calls[call_label]
        ws.append([f"Call {call_label}", c.get("UP",0), c.get("FLAT",0), c.get("DOWN",0), sum(c.values())])

    h_totals_up   = sum(human_calls[c].get("UP",0) for c in human_calls)
    h_totals_flat = sum(human_calls[c].get("FLAT",0) for c in human_calls)
    h_totals_down = sum(human_calls[c].get("DOWN",0) for c in human_calls)
    ws.append(["Col total", h_totals_up, h_totals_flat, h_totals_down, h_totals_up+h_totals_flat+h_totals_down])
    ws.append(["Source: human_decisions_export_2026-08-12.csv + returns_matrix.csv", "", "", "", ""])

    # ── (c) Human vs model agreement (from kappa CSV) ────────────────────────
    add_header("(c) Human vs model agreement — N=171 paired events. Source: kappa_near_independence.csv")
    ws.append(["Human call \\ LLM call", "LLM BUY", "LLM HOLD", "LLM SELL", "Row total"])
    ws[ws.max_row][0].font = bold

    kappa_rows = []
    with open(KAPPA_CSV) as f:
        content = f.read()
    lines = [l for l in content.splitlines() if not l.startswith("#")]
    reader = csv.DictReader(lines)
    for row in reader:
        kappa_rows.append(row)

    conf_start = False
    for row in kappa_rows:
        if row.get("metric","").startswith("Human_"):
            conf_start = True
        if conf_start:
            label = row.get("metric","")
            if label.startswith("Human_"):
                arm = label.replace("Human_","Human ").replace("_"," ")
                vals = [row.get("LLM_BUY",""), row.get("LLM_HOLD",""), row.get("LLM_SELL",""), row.get("row_total","")]
                ws.append([arm] + vals)

    kappa_map = {r["metric"]: r["value"] for r in kappa_rows if r.get("metric") and r.get("value")}
    ws.append([])
    ws.append(["Kappa", kappa_map.get("cohens_kappa","0.109")])
    ws.append(["90% CI", f"[{kappa_map.get('bootstrap_90ci_low','0.030')}, {kappa_map.get('bootstrap_90ci_high','0.190')}]"])
    ws.append(["Observed agreement", kappa_map.get("observed_agreement","")])
    ws.append(["Expected agreement", kappa_map.get("expected_agreement","")])

    ws.freeze_panes = "A2"
    for i in range(1, 9):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = 30
    return ws


def build_anchor_audit_sheet(wb, rm):
    """Build Anchor_Audit sheet."""
    headers = ["document_id", "ticker", "release_date", "entry_date", "timing",
               "entry_shifted", "verification_method", "timing_excluded", "notes"]
    rows_data = []
    for did, r in sorted(rm.items()):
        entry_date = r["entry_date"]
        release_date = r["report_date"]
        if entry_date < release_date:
            timing = "pre_market"
        elif entry_date == release_date:
            timing = "after_hours"
        else:
            timing = "unknown"
        shifted = entry_date != release_date
        rows_data.append([
            did, r["ticker"], release_date, entry_date, timing,
            "YES" if shifted else "NO",
            "returns_matrix.csv (corrected 2026-08-12)",
            r.get("timing_excluded","NO"),
            "",
        ])
    make_sheet(wb, "Anchor_Audit", headers, rows_data)


def build_item_c_sheet(wb):
    """Build Item_C_Ablation sheet — data rows only from committed CSVs."""
    ws = wb.create_sheet(title="Item_C_Ablation")
    bold = BOLD

    # Per-arm table
    headers_cost = ["event_set", "arm", "n_calls", "n_graded", "n_correct",
                    "accuracy_hold_excluded", "mean_tokens_per_call", "total_cost_usd",
                    "cost_per_call_usd", "cost_per_correct_usd"]
    ws.append(["PER-ARM ACCURACY AND COST"])
    ws[ws.max_row][0].font = bold
    ws.append(headers_cost)
    for cell in ws[ws.max_row]:
        cell.font = bold

    abl_cost = load_csv_data(ABLATION_COST)
    for row in abl_cost:
        ws.append([row.get(h,"") for h in headers_cost])

    ws.append([])
    ws.append(["Note: figures from section_ablation_cost_per_correct.csv DATA ROWS ONLY. "
               "four_arm_n119 = events where all 4 arms scored. all_n200 = full scored set for full_bundle and press_release."])

    # Paired diffs
    ws.append([])
    ws.append(["PAIRED DIFFERENCES VS FULL BUNDLE"])
    ws[ws.max_row][0].font = bold
    headers_paired = ["comparison","set","method","n_paired","point_diff","ci_low","ci_high","p_value","metric","note"]
    ws.append(headers_paired)
    for cell in ws[ws.max_row]:
        cell.font = bold

    abl_paired = load_csv_data(ABLATION_PAIRED)
    for row in abl_paired:
        ws.append([row.get(h,"") for h in headers_paired])

    ws.append([])
    ws.append(["Paired MDE ~±12pp. 6.8pp gap at p=0.055 on inclusive test (press_release vs full_bundle). "
               "Strict test (both graded, n=45): exactly 0.0 — when both arms commit, they commit identically."])
    ws.append(["Prepared remarks agrees with full bundle on 85.7% of signals at ~10k tokens."])

    ws.freeze_panes = "A2"
    for i in range(1, 11):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = 20
    return ws


def build_walkforward_sheet(wb):
    """Build Walk_Forward sheet."""
    with open(WALKFORWARD_JSON) as f:
        wf = json.load(f)

    ws = wb.create_sheet(title="Walk_Forward")
    bold = BOLD

    def section(text):
        ws.append([])
        ws.append([text])
        ws[ws.max_row][0].font = bold
        ws.append([])

    section("DEGENERACY FINDING")
    rw = wf.get("rolling_walkforward", {})
    ws.append(["n_windows", rw.get("n_windows","")])
    ws.append(["min_trades_floor", rw.get("min_trades_floor","")])
    ws.append(["objectives_tried", str(rw.get("objectives_tried",""))])
    ws.append(["degeneracy_finding", str(rw.get("degeneracy_finding",""))])
    for obj in ["mean_net_objective","accuracy_objective"]:
        o = rw.get(obj, {})
        ws.append([obj + "_n_degenerate", o.get("n_degenerate_windows",""), o.get("note","")])

    section("IN-SAMPLE DEPLOYED THRESHOLDS")
    isd = wf.get("in_sample_deployed", {})
    for k, v in isd.items():
        ws.append([k, v])

    section("DEV / EVAL SPLIT")
    ws.append(["Split rule", "Sort 233 clean events by release_date (returns_matrix.csv), earliest 20% (47) = dev, remaining 80% (186) = eval."])
    ws.append(["Eval: accuracy", "68.0% (51/75)", "floor=54.6% (65/119 graded)", "margin +13.4pp", "p=0.013", "MDE ±14.4pp"])
    ws.append(["Dev:  accuracy", "55.0% (11/20)", "floor=53.6%", "margin +1.4pp", "NOT significant", "MDE ±27.8pp — underpowered"])
    ws.append(["Source", "frontier_table.csv, surviving_findings.md"])
    ws.append(["NOT out-of-sample", "Dev/eval split applied post hoc to data thresholds were fitted on. Eval 68.0% is in-sample on a subset."])

    section("CROSS-ISSUER GENERALISATION (RECONSTRUCTED)")
    gu = wf.get("genuinely_unseen", {})
    ps = gu.get("post_sweep", {})
    ins = gu.get("in_sweep", {})
    ws.append(["note", gu.get("note","")])
    ws.append([])
    ws.append(["subset", "n_events", "n_trades", "n_graded", "n_correct", "accuracy", "mean_net_pct", "floor", "margin", "p", "mde"])
    for label, d in [("in_sweep", ins), ("post_sweep", ps)]:
        floor = d.get("majority_direction_floor","")
        ws.append([label,
                   gu.get(f"n_events_{'in' if 'in' in label else 'post'}_sweep",""),
                   d.get("n_trades",""), d.get("n_graded",""), d.get("n_correct",""),
                   d.get("accuracy",""), d.get("mean_net_pct",""),
                   floor, d.get("margin_vs_floor_pp",""),
                   d.get("binomial_p",""), d.get("mde_pp","")])
    ws.append([])
    ws.append(["CAVEATS (three)"])
    ws.append(["1. RECONSTRUCTED", "Sweep membership inferred from issuer ordering and count match (n=161). 453 single-swap alternatives also give n=161."])
    ws.append(["2. ISSUER TRANSFER ONLY", "Both subsets overlap in time. Tests whether thresholds transfer to new companies, not new time periods. Does NOT corroborate temporal generalisation."])
    ws.append(["3. NOT SIGNIFICANT", f"p={ps.get('binomial_p','')} does not clear 0.05. MDE=±{ps.get('mde_pp','')}pp — subset underpowered. Report as directionally consistent, not confirmed."])

    ws.freeze_panes = "A2"
    ws.column_dimensions["A"].width = 35
    ws.column_dimensions["B"].width = 60
    for i in range(3, 12):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = 15
    return ws


# ─────────────────────────────────────────────────────────────────────────────
# Write corrected workbooks
# ─────────────────────────────────────────────────────────────────────────────

BANNER_TEXT = ("HUMAN ARM FROZEN. Prices and derived accuracy on the original report_date basis, "
               "superseded 2026-08-12. Net P&L in column U is a formula deriving from those prices "
               "and therefore remains on the superseded basis by design. Corrected figures are in "
               "Corrections_Log and Corrected_Metrics.")


def write_llm_sheet(ws, llm_rows):
    """Write corrected values to LLM_Data_Entry. All literals."""
    for rec in llm_rows:
        r = rec["row_index"]
        # Map rec fields to columns
        # G: sentiment_score, H: decision, J: document_date, K: closing_date,
        # L: prior_close, M: opening_date, N: next_day_open, O: token_cost,
        # P: actual_pct_change, Q: actual_direction, R: prediction_correct,
        # S: position, T: net_pnl, U: notes

        def to_date(s):
            if not s or s == "":
                return None
            try:
                return datetime.datetime.strptime(str(s), "%Y-%m-%d")
            except Exception:
                return None

        ws.cell(r, L_SCORE).value  = rec["sentiment_score"]
        ws.cell(r, L_DEC).value    = rec["decision"]
        ws.cell(r, L_DOC).value    = to_date(rec["document_date"])
        ws.cell(r, L_K).value      = to_date(rec["closing_date"])
        ws.cell(r, L_L).value      = float(rec["prior_close"]) if rec["prior_close"] != "" else None
        ws.cell(r, L_M).value      = to_date(rec["opening_date"])
        ws.cell(r, L_N).value      = float(rec["next_day_open"]) if rec["next_day_open"] else None
        ws.cell(r, L_O_TOK).value  = float(rec["token_cost"]) if rec["token_cost"] else None
        ws.cell(r, L_P_PCT).value  = float(rec["actual_pct_change"]) / 100 if rec["actual_pct_change"] != "" else None
        ws.cell(r, L_Q_DIR).value  = rec["actual_direction"]
        ws.cell(r, L_R_COR).value  = rec["prediction_correct"]
        ws.cell(r, L_S_POS).value  = int(rec["position"])
        ws.cell(r, L_T_NET).value  = float(rec["net_pnl"])
        if rec.get("notes"):
            ws.cell(r, L_U_NOT).value = rec["notes"]


def build_workbook(path, human_rows, llm_rows, rm, excluded, correct_human):
    """Build one output workbook. correct_human=True writes corrected prices."""
    print(f"\n  Building {path.name} ...")
    shutil.copy(SOURCE_WB, path)
    wb = openpyxl.load_workbook(path, data_only=False)
    ws_h = wb["Human_Data_Entry"]
    ws_l = wb["LLM_Data_Entry"]

    # ── A1 banner on Human_Data_Entry ────────────────────────────────────────
    orig_font = ws_h.cell(1, 1).font.copy() if ws_h.cell(1, 1).font else Font()
    ws_h.cell(1, 1).value = BANNER_TEXT
    # Keep bold if already bold
    ws_h.cell(1, 1).font = Font(bold=True, name=orig_font.name, size=orig_font.size)

    # ── Human prices (CORRECTED workbook only) ───────────────────────────────
    n_written = n_skipped = 0
    alignment_failures = []

    if correct_human:
        # Build row_index → correction record
        corr_map = {r["row_index"]: r for r in human_rows}
        for row_idx in range(3, ws_h.max_row + 1):
            co = ws_h.cell(row_idx, H_CO).value
            if not co:
                continue
            yr  = ws_h.cell(row_idx, H_YR).value
            qtr = ws_h.cell(row_idx, H_QTR).value
            sheet_key = f"{co}|{yr}|{qtr}"

            rec = corr_map.get(row_idx)
            if rec is None:
                continue

            # Row alignment assertion
            rec_key = rec["event_key"]
            if sheet_key != rec_key:
                alignment_failures.append(
                    f"Row {row_idx}: sheet key={sheet_key!r}, export key={rec_key!r}"
                )
                # Do not write; continue
                continue

            if rec["correctable"] != "TRUE":
                n_skipped += 1
                continue

            # Write M, N, O, P only
            try:
                ws_h.cell(row_idx, H_M).value = datetime.datetime.strptime(rec["prior_closing_date"], "%Y-%m-%d")
                ws_h.cell(row_idx, H_N).value = float(rec["prior_close"])
                ws_h.cell(row_idx, H_O).value = datetime.datetime.strptime(rec["next_opening_date"], "%Y-%m-%d")
                ws_h.cell(row_idx, H_P).value = float(rec["next_day_open"])
                n_written += 1
            except Exception as e:
                alignment_failures.append(f"Row {row_idx} write error: {e}")
    else:
        # Locked workbook: verify alignment only, write nothing to Human
        corr_map = {r["row_index"]: r for r in human_rows}
        for row_idx in range(3, ws_h.max_row + 1):
            co = ws_h.cell(row_idx, H_CO).value
            if not co:
                continue
            yr  = ws_h.cell(row_idx, H_YR).value
            qtr = ws_h.cell(row_idx, H_QTR).value
            sheet_key = f"{co}|{yr}|{qtr}"
            rec = corr_map.get(row_idx)
            if rec and sheet_key != rec["event_key"]:
                alignment_failures.append(
                    f"Row {row_idx}: sheet={sheet_key!r} export={rec['event_key']!r}"
                )

    # ── LLM sheet ────────────────────────────────────────────────────────────
    write_llm_sheet(ws_l, llm_rows)

    # ── New sheets ───────────────────────────────────────────────────────────
    build_corrections_log_sheet(wb, human_rows, llm_rows)

    # Corrected_Metrics
    metrics_rows_data = []
    metrics_headers = [
        "metric_name","value","unit","n","denominator_definition","event_set","convention",
        "ci_low","ci_high","ci_level","p_value","mde","test_used","threshold_dependent",
        "source_file","notes",
    ]
    metrics_rows = load_csv_data(OUT_METRICS)
    for mr in metrics_rows:
        metrics_rows_data.append([mr.get(h,"") for h in metrics_headers])
    make_sheet(wb, "Corrected_Metrics", metrics_headers, metrics_rows_data)

    build_findings_sheet(wb)
    build_confusion_matrices_sheet(wb, rm, excluded)
    build_anchor_audit_sheet(wb, rm)
    build_item_c_sheet(wb)
    build_walkforward_sheet(wb)

    wb.save(path)
    wb.close()

    print(f"    Alignment checks: {len(alignment_failures)} failures out of 420 rows")
    if alignment_failures:
        print("    ALIGNMENT FAILURES (first 5):")
        for f in alignment_failures[:5]:
            print(f"      {f}")
    if correct_human:
        print(f"    Human rows written: {n_written}, skipped (not correctable): {n_skipped}")
    return alignment_failures, n_written if correct_human else 0, n_skipped if correct_human else 0


# ─────────────────────────────────────────────────────────────────────────────
# Verification
# ─────────────────────────────────────────────────────────────────────────────

def verify_workbook(path):
    """Verify structural properties of an output workbook."""
    wb = openpyxl.load_workbook(path, data_only=False)
    results = {}
    results["sheet_count"] = len(wb.sheetnames)
    results["sheets"] = wb.sheetnames
    charts_ws = wb["Charts"] if "Charts" in wb.sheetnames else None
    eff_ws    = wb["Efficiency"] if "Efficiency" in wb.sheetnames else None
    results["charts_Charts"] = len(charts_ws._charts) if charts_ws else "MISSING"
    results["charts_Efficiency"] = len(eff_ws._charts) if eff_ws else "MISSING"
    # Tables
    total_tables = sum(len(wb[s].tables) for s in wb.sheetnames)
    results["table_count"] = total_tables
    # Human S3 formula
    ws_h = wb["Human_Data_Entry"]
    s3 = ws_h.cell(3, 19).value  # col S = 19
    results["s3_formula_check"] = str(s3)[:40] if s3 else "MISSING"
    results["s3_starts_with_IF_OR_N3"] = str(s3).startswith("=IF(OR(N3") if s3 else False
    wb.close()
    return results


# ─────────────────────────────────────────────────────────────────────────────
# Main
# ─────────────────────────────────────────────────────────────────────────────

def main():
    print("Loading source data...")
    wb_src = openpyxl.load_workbook(SOURCE_WB, data_only=False)
    rm       = load_returns_matrix()
    cal      = load_calibration()
    costs    = load_cost_ledger()
    excluded = load_exclusion_set()
    ablation = load_ablation_by_event()
    print(f"  returns_matrix: {len(rm)} events")
    print(f"  calibration: {len(cal)} events")
    print(f"  excluded: {len(excluded)} events")

    print("\n=== FILE 1: Human prices ===")
    human_rows, assert_failures_1 = build_human_prices(wb_src, rm, excluded)

    print("\n=== FILE 2: LLM CSV ===")
    llm_rows = build_llm_csv(wb_src, rm, cal, costs, excluded, ablation)

    print("\n=== FILE 3: Metrics CSV ===")
    build_metrics_csv()

    print("\n=== FILE 4: README ===")
    build_readme()

    wb_src.close()

    # ── Overwrite guard ───────────────────────────────────────────────────────
    # The CORRECTED workbook accumulates hand-applied patches (fix_workbooks_v*.py,
    # KHC sentinel fix, Accuracy_Conventions edits) that build_workbook.py cannot
    # reconstruct from source data.  Silently overwriting it discards those patches.
    # Pass --force to override explicitly.
    force = "--force" in sys.argv
    if OUT_CORRECTED.exists() and not force:
        print(
            f"\nERROR: {OUT_CORRECTED.name} already exists.\n"
            f"  build_workbook.py recreates this file from {SOURCE_WB.name},\n"
            f"  which will discard any hand-applied patches committed on top of it\n"
            f"  (Accuracy_Conventions edits, KHC fix, In Clean Universe columns, etc.).\n"
            f"  If you genuinely want to rebuild from scratch, pass --force.\n"
            f"  If you only want to rebuild the CSV exports, omit --no-workbooks (not yet implemented).\n"
            f"  Aborting without writing any workbook."
        )
        sys.exit(1)

    print("\n=== WORKBOOK 1: LOCKED ===")
    wb_src2 = openpyxl.load_workbook(SOURCE_WB, data_only=False)
    llm_rows_dicts = [dict(row) for row in llm_rows]

    fails_locked, written_locked, skipped_locked = build_workbook(
        OUT_LOCKED, human_rows, llm_rows, rm, excluded, correct_human=False
    )

    print("\n=== WORKBOOK 2: CORRECTED ===")
    fails_corrected, written_corrected, skipped_corrected = build_workbook(
        OUT_CORRECTED, human_rows, llm_rows, rm, excluded, correct_human=True
    )

    print("\n=== VERIFICATION ===")
    for path, label in [(OUT_LOCKED, "LOCKED"), (OUT_CORRECTED, "CORRECTED")]:
        v = verify_workbook(path)
        print(f"\n  {label}:")
        print(f"    Sheet count: {v['sheet_count']} (expected 14: 7 original + 7 new)")
        print(f"    Sheets: {v['sheets']}")
        print(f"    Charts/Charts sheet: {v['charts_Charts']} (expected 5)")
        print(f"    Charts/Efficiency sheet: {v['charts_Efficiency']} (expected 4)")
        print(f"    Table count: {v['table_count']} (expected ~12)")
        print(f"    H_S3 formula check: {v['s3_formula_check']!r}")
        print(f"    S3 starts with =IF(OR(N3: {v['s3_starts_with_IF_OR_N3']}")

    print("\n=== ROW ALIGNMENT SUMMARY ===")
    print(f"  LOCKED: {len(fails_locked)} alignment failures on 420 rows")
    print(f"  CORRECTED: {len(fails_corrected)} alignment failures on 420 rows; "
          f"{written_corrected} rows written, {skipped_corrected} skipped (not correctable)")

    if fails_locked or fails_corrected:
        print("  WARNING: alignment failures detected — do not circulate these files")
        all_failures = set(fails_locked) | set(fails_corrected)
        for f in list(all_failures)[:10]:
            print(f"    {f}")
    else:
        print("  Assertion PASSED: all 420 row alignment checks passed in both workbooks.")

    print("\n=== NOTE FOR USER ===")
    print("  Open each file in Excel and press Ctrl+Alt+F9 (force full recalculation)")
    print("  before circulating. LibreOffice will show #NAME? for XLOOKUP — this is")
    print("  expected and does not indicate file corruption.")
    print("\nDone.")
    return 0


if __name__ == "__main__":
    sys.exit(main())
